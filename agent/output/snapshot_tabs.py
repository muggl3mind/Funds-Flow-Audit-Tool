"""
Add PDF snapshot tabs to the annotated workpaper.

For each matched document, renders the first page of the PDF as an image
and inserts it into a dedicated Excel tab named "FF01 · Vendor Name".
Requires pdf2image + Pillow + poppler (optional — skips gracefully if missing).
"""
from __future__ import annotations

import io
import re
from pathlib import Path

from openpyxl.styles import Alignment, Font, PatternFill


def add_snapshots(wb, deal_dir: Path, index: dict) -> None:
    """Add one sheet per matched document with a rendered PDF image."""
    try:
        from pdf2image import convert_from_path
        from PIL import Image as PILImage
    except ImportError:
        print("  pdf2image / Pillow not available — skipping snapshots")
        return

    from openpyxl.drawing.image import Image as XLImage

    docs_dir = deal_dir / "documents"
    seen: set[str] = set()

    for item in index["line_items"]:
        doc_file = item.get("document_file")
        ff_ref   = item.get("ff_ref")
        if not doc_file or not ff_ref or item.get("status") == "MISSING":
            continue

        pdf_path = docs_dir / doc_file
        if not pdf_path.exists():
            print(f"  snapshot: file not found — {doc_file}")
            continue

        if doc_file not in seen:
            vendor_short = re.sub(r'[/\\?*:\[\]]', '-',
                                  (item.get("document_vendor") or doc_file)[:20])
            tab_name = f"{ff_ref} · {vendor_short}"[:31]

            ws = wb.create_sheet(tab_name)
            ws.column_dimensions["A"].width = 108
            ws.sheet_view.showGridLines = False

            ws.merge_cells("A1:B1")
            t = ws.cell(row=1, column=1,
                        value=f"{ff_ref}  |  {item.get('document_vendor', '')}  |  {doc_file}")
            t.font      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            t.fill      = PatternFill("solid", fgColor="1F3864")
            t.alignment = Alignment(horizontal="left", indent=1)
            ws.row_dimensions[1].height = 22

            pages = convert_from_path(str(pdf_path), dpi=130, first_page=1, last_page=1)
            img   = pages[0]
            scale = 800 / img.width
            new_h = int(img.height * scale)
            img   = img.resize((800, new_h), PILImage.LANCZOS)

            buf = io.BytesIO()
            img.save(buf, format="PNG")
            buf.seek(0)

            xl_img        = XLImage(buf)
            xl_img.anchor = "A2"
            ws.add_image(xl_img)
            ws.row_dimensions[2].height = new_h * 0.74

            seen.add(doc_file)
            print(f"  snapshot: {tab_name}")
