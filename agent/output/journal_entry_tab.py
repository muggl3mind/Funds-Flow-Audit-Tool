"""
Build the Journal Entry tab inside the annotated workpaper.

Creates two fully separate journal entries (Fund I and Fund II), each with
debits, credits (Wire Payable vs Accrued Liabilities), and a balance check.
Amounts are formula-linked to the Buyer Expenses tab.
"""
from __future__ import annotations

from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Font

from agent.output.styles import (
    BOLD_FONT, CELL_FONT, HEADER_FILL, HEADER_FONT,
    STATUS_COLORS, THIN_BORDER, MED_BOTTOM, USD_FMT, CHK_FMT,
)
from agent.output.workpaper_annotator import load_credit_accounts

CREDIT_ACCOUNTS = load_credit_accounts()

COL_HEADERS = [
    "Entry ID", "Date", "Account Code", "Account Name",
    "Description", "FF Ref", "Debit", "Credit", "Entry Type",
]
COL_WIDTHS = [10, 12, 14, 42, 55, 8, 14, 14, 10]


def _cell_ref(sheet: str, col: str, row: int) -> str:
    return f"'{sheet}'!{col}{row}"


def _hcell(ws, r, c, val):
    cell = ws.cell(r, c, val)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER
    return cell


def _dcell(ws, r, c, val, *, bold=False, align="left", fill=None, fmt=None):
    cell = ws.cell(r, c, val)
    cell.font = BOLD_FONT if bold else CELL_FONT
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    return cell


def build(wb, index: dict, je_row_info: list[dict]) -> None:
    """Insert a 'Journal Entry' tab into the workbook."""
    closing_date = index["closing_date"]
    deal         = index["deal"]

    ws = wb.create_sheet("Journal Entry")

    for c, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # Title banner
    ws.merge_cells("A1:I1")
    title = ws["A1"]
    title.value = (
        f"JOURNAL ENTRY  |  {deal}  |  Closing: {closing_date}  |  "
        "Fund I: FF-JE-001-FI   Fund II: FF-JE-001-FII"
    )
    title.font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    title.fill = HEADER_FILL
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:I2")
    note = ws["A2"]
    note.value = (
        "⚠  DRAFT — Debit/Credit amounts are formula-linked to 'Buyer Expenses' tab. "
        "Review with fund accounting before posting. Fund I and Fund II post separately."
    )
    note.font = Font(name="Calibri", italic=True, color="C00000", size=10)
    note.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # Pre-compute credit formula parts
    def _get_ref(rec: dict, col_key: str) -> str:
        if rec["sheet"] and rec[col_key]:
            return _cell_ref(rec["sheet"], rec[col_key], rec["row"])
        item_obj = next(
            (i for i in index["line_items"] if i["description"] == rec["description"]), None)
        if item_obj:
            return str(item_obj["fund_i_amount"] if col_key == "fi_col" else item_obj["fund_ii_amount"])
        return "0"

    wp_parts: dict[str, list[str]] = {"fi_col": [], "fii_col": []}
    ac_parts: dict[str, list[str]] = {"fi_col": [], "fii_col": []}
    wp_descs: list[str] = []
    ac_descs: list[str] = []

    for rec in je_row_info:
        fi_ref  = _get_ref(rec, "fi_col")
        fii_ref = _get_ref(rec, "fii_col")
        ff_ref  = rec["ff_ref"] or "—"
        status  = rec["status"]
        ff_amt  = rec["ff_amount"]
        doc_amt = rec["doc_amount"]

        if status in ("MATCHED", "CUMULATIVE"):
            wp_parts["fi_col"].append(fi_ref)
            wp_parts["fii_col"].append(fii_ref)
            wp_descs.append(f"{ff_ref} {rec['description']}")
        elif status == "PARTIAL" and ff_amt:
            shortfall = ff_amt - doc_amt
            wp_parts["fi_col"].append(f"{fi_ref}*{doc_amt}/{ff_amt}")
            wp_parts["fii_col"].append(f"{fii_ref}*{doc_amt}/{ff_amt}")
            ac_parts["fi_col"].append(f"{fi_ref}*{shortfall}/{ff_amt}")
            ac_parts["fii_col"].append(f"{fii_ref}*{shortfall}/{ff_amt}")
            wp_descs.append(f"{ff_ref} {rec['description']} (${doc_amt:,.0f} of ${ff_amt:,.0f})")
            ac_descs.append(f"{ff_ref} {rec['description']} (shortfall ${shortfall:,.0f})")
        elif status == "MISSING":
            ac_parts["fi_col"].append(fi_ref)
            ac_parts["fii_col"].append(fii_ref)
            ac_descs.append(f"— {rec['description']} (no invoice)")

    def _cr_formula(parts: list[str]) -> str | int:
        return ("=" + "+".join(parts)) if parts else 0

    credit_fill_wp = PatternFill("solid", fgColor="C6EFCE")
    credit_fill_ac = PatternFill("solid", fgColor="FFEB9C")

    def _write_fund_section(start_row: int, fund_label: str,
                            je_ref: str, col_key: str) -> int:
        r = start_row

        # Fund banner
        ws.merge_cells(f"A{r}:I{r}")
        banner = ws[f"A{r}"]
        banner.value = f"{fund_label}  —  {je_ref}"
        banner.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        banner.fill = PatternFill("solid", fgColor="2E4A7A")
        banner.alignment = Alignment(horizontal="left", indent=1, vertical="center")
        ws.row_dimensions[r].height = 20
        r += 1

        # Debit section
        ws.merge_cells(f"A{r}:I{r}")
        dl = ws[f"A{r}"]
        dl.value = "DEBIT LINES"
        dl.font = BOLD_FONT
        dl.fill = PatternFill("solid", fgColor="DDEBF7")
        dl.alignment = Alignment(horizontal="left", indent=1, vertical="center")
        r += 1

        for c, h in enumerate(COL_HEADERS, start=1):
            _hcell(ws, r, c, h)
        ws.row_dimensions[r].height = 18
        r += 1

        debit_rows: list[int] = []
        for rec in je_row_info:
            ref      = _get_ref(rec, col_key)
            formula  = f"={ref}"
            ff_ref   = rec["ff_ref"] or "—"
            desc_str = f"{ff_ref}  {rec['description']}  —  {rec['vendor']}"
            status   = rec["status"]
            fill     = PatternFill("solid", fgColor=STATUS_COLORS.get(status, "FFFFFF"))

            _dcell(ws, r, 1, je_ref,           align="center")
            _dcell(ws, r, 2, closing_date,     align="center")
            _dcell(ws, r, 3, rec["acct_code"], align="center")
            _dcell(ws, r, 4, rec["acct_name"])
            _dcell(ws, r, 5, desc_str)
            _dcell(ws, r, 6, ff_ref,           align="center", fill=fill)
            _dcell(ws, r, 7, formula,          align="right",  fmt=USD_FMT)
            _dcell(ws, r, 8, 0,                align="right",  fmt=USD_FMT)
            _dcell(ws, r, 9, "debit",          align="center")
            debit_rows.append(r)
            r += 1

        dr_sum_f = "=SUM(" + ",".join(f"G{dr}" for dr in debit_rows) + ")"
        ws.cell(r, 6, "Total Debits").font = BOLD_FONT
        ws.cell(r, 6).alignment = Alignment(horizontal="right")
        td = ws.cell(r, 7, dr_sum_f)
        td.font = BOLD_FONT
        td.number_format = USD_FMT
        td.alignment = Alignment(horizontal="right")
        td.border = MED_BOTTOM
        r += 2

        # Credit section
        ws.merge_cells(f"A{r}:I{r}")
        cl = ws[f"A{r}"]
        cl.value = "CREDIT LINES"
        cl.font = BOLD_FONT
        cl.fill = PatternFill("solid", fgColor="C6EFCE")
        cl.alignment = Alignment(horizontal="left", indent=1, vertical="center")
        r += 1

        for c, h in enumerate(COL_HEADERS, start=1):
            _hcell(ws, r, c, h)
        ws.row_dimensions[r].height = 18
        r += 1

        credit_rows: list[int] = []
        wp_f = _cr_formula(wp_parts[col_key])
        ac_f = _cr_formula(ac_parts[col_key])
        wp_basis = "; ".join(wp_descs[:3]) + ("…" if len(wp_descs) > 3 else "")

        wp_code, wp_name = CREDIT_ACCOUNTS.get("wire_payable",        ("2100", "Wire Payable"))
        ac_code, ac_name = CREDIT_ACCOUNTS.get("accrued_liabilities", ("2200", "Accrued Liabilities"))

        for acct_name, acct_code, cr_formula, fill, basis in [
            (wp_name, wp_code, wp_f, credit_fill_wp, wp_basis),
            (ac_name, ac_code, ac_f, credit_fill_ac, "; ".join(ac_descs)),
        ]:
            _dcell(ws, r, 1, je_ref,      align="center")
            _dcell(ws, r, 2, closing_date, align="center")
            _dcell(ws, r, 3, acct_code,   align="center")
            _dcell(ws, r, 4, acct_name,   bold=True, fill=fill)
            _dcell(ws, r, 5, basis,       fill=fill)
            _dcell(ws, r, 6, "",          fill=fill)
            _dcell(ws, r, 7, 0,           align="right", fmt=USD_FMT)
            _dcell(ws, r, 8, cr_formula,  align="right", fmt=USD_FMT, fill=fill)
            _dcell(ws, r, 9, "credit",    align="center")
            credit_rows.append(r)
            r += 1

        cr_sum_f = "=SUM(" + ",".join(f"H{cr}" for cr in credit_rows) + ")"
        ws.cell(r, 6, "Total Credits").font = BOLD_FONT
        ws.cell(r, 6).alignment = Alignment(horizontal="right")
        tc = ws.cell(r, 9, cr_sum_f)
        tc.font = BOLD_FONT
        tc.number_format = USD_FMT
        tc.alignment = Alignment(horizontal="right")
        tc.border = MED_BOTTOM
        r += 2

        # Balance check
        dr_cells = ",".join(f"G{dr}" for dr in debit_rows)
        cr_cells = ",".join(f"H{cr}" for cr in credit_rows)
        ws.merge_cells(f"A{r}:F{r}")
        bc_label = ws[f"A{r}"]
        bc_label.value = f"Balance Check — {fund_label}  (Debits − Credits = 0)"
        bc_label.font = BOLD_FONT
        bc_label.alignment = Alignment(horizontal="right", vertical="center")
        bc = ws.cell(r, 7, f"=SUM({dr_cells})-SUM({cr_cells})")
        bc.font = BOLD_FONT
        bc.number_format = CHK_FMT
        bc.alignment = Alignment(horizontal="right")
        bc.border = THIN_BORDER
        r += 3

        return r

    next_row = _write_fund_section(4, "Fund I",  "FF-JE-001-FI",  "fi_col")
    _write_fund_section(next_row, "Fund II", "FF-JE-001-FII", "fii_col")

    ws.freeze_panes = "A4"
