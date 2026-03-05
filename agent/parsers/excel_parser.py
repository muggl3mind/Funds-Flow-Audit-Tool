"""
Stage 1: Parse a multi-tab Excel funds flow into raw structural objects.

No semantic interpretation here — just faithful extraction of cells, merged
regions, hidden rows, and subtotal hints. All semantic decisions happen in
the normalizer (Stage 2).
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.utils import get_column_letter

from agent.utils.amount_utils import infer_unit_scale, parse_amount


@dataclass
class RawCell:
    row: int
    col: int
    col_letter: str
    value: Any                 # str, int, float, None — already numeric where possible
    number_format: str
    unit_scale: float          # derived from number_format
    parsed_amount: Optional[float]  # set if cell looks numeric
    is_merged_child: bool      # True if this cell inherits value from a merge parent


@dataclass
class RawRow:
    row_idx: int
    cells: list[RawCell]
    is_hidden: bool
    is_subtotal: bool          # heuristic: description text contains total/subtotal keyword
    is_section_header: bool    # heuristic: has text but no numeric cells
    as_text: str               # quick single-line representation for LLM prompts


@dataclass
class RawSheet:
    name: str
    raw_rows: list[RawRow]
    merged_regions: list[str]  # e.g. ["A1:D1", "A3:B3"]
    max_col: int


def parse_workbook(path: Path) -> list[RawSheet]:
    """Open an xlsx/xls file and return one RawSheet per worksheet."""
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=False)
    sheets: list[RawSheet] = []

    for ws in wb.worksheets:
        merged = [str(r) for r in ws.merged_cells.ranges]

        # Build a map: (row, col) → merged-parent value
        merge_value_map: dict[tuple[int, int], Any] = {}
        for region in ws.merged_cells.ranges:
            parent = ws.cell(region.min_row, region.min_col)
            parent_val = parent.value
            for r in range(region.min_row, region.max_row + 1):
                for c in range(region.min_col, region.max_col + 1):
                    if (r, c) != (region.min_row, region.min_col):
                        merge_value_map[(r, c)] = parent_val

        raw_rows: list[RawRow] = []
        max_col = ws.max_column or 1

        for row_idx in range(1, (ws.max_row or 1) + 1):
            rd = ws.row_dimensions.get(row_idx)
            is_hidden = bool(rd and rd.hidden)

            raw_cells: list[RawCell] = []
            has_text = False
            has_number = False

            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row_idx, col_idx)
                is_child = (row_idx, col_idx) in merge_value_map
                value = merge_value_map[(row_idx, col_idx)] if is_child else cell.value
                nf = (cell.number_format or "") if not is_child else ""
                scale = infer_unit_scale(nf)
                parsed = parse_amount(value, scale) if value is not None else None

                if isinstance(value, str) and value.strip():
                    has_text = True
                if parsed is not None:
                    has_number = True

                raw_cells.append(RawCell(
                    row=row_idx,
                    col=col_idx,
                    col_letter=get_column_letter(col_idx),
                    value=value,
                    number_format=nf,
                    unit_scale=scale,
                    parsed_amount=parsed,
                    is_merged_child=is_child,
                ))

            # Subtotal heuristic: first text cell contains "total" / "subtotal"
            first_text = next(
                (c.value for c in raw_cells if isinstance(c.value, str) and c.value.strip()), ""
            )
            is_subtotal = bool(first_text and
                               any(kw in first_text.lower() for kw in
                                   ("total", "subtotal", "sub-total", "grand total", "sum")))

            # Section header heuristic: has text, no numbers, usually bold (we can't read bold
            # reliably in data_only mode so we rely on "no amounts" heuristic)
            is_section_header = has_text and not has_number and not is_subtotal

            as_text = _row_as_text(row_idx, raw_cells, is_subtotal, is_hidden)

            raw_rows.append(RawRow(
                row_idx=row_idx,
                cells=raw_cells,
                is_hidden=is_hidden,
                is_subtotal=is_subtotal,
                is_section_header=is_section_header,
                as_text=as_text,
            ))

        sheets.append(RawSheet(
            name=ws.title,
            raw_rows=raw_rows,
            merged_regions=merged,
            max_col=max_col,
        ))

    wb.close()
    return sheets


def sheet_to_prompt_text(sheet: RawSheet, max_rows: int = 80) -> str:
    """
    Render a RawSheet as a compact text table suitable for an LLM prompt.
    Skips completely empty rows. Truncates at max_rows.
    """
    lines = [f"=== Sheet: {sheet.name!r} ==="]
    count = 0
    for row in sheet.raw_rows:
        if all(c.value is None for c in row.cells):
            continue                          # skip blank rows
        lines.append(row.as_text)
        count += 1
        if count >= max_rows:
            lines.append(f"... (truncated at {max_rows} rows)")
            break
    return "\n".join(lines)


def _row_as_text(row_idx: int, cells: list[RawCell],
                 is_subtotal: bool, is_hidden: bool) -> str:
    tags = []
    if is_subtotal:
        tags.append("SUBTOTAL")
    if is_hidden:
        tags.append("HIDDEN")
    tag_str = f"[{', '.join(tags)}] " if tags else ""

    vals = []
    for c in cells:
        if c.value is None:
            vals.append("")
        elif c.parsed_amount is not None and not isinstance(c.value, str):
            vals.append(f"${c.parsed_amount:,.0f}")
        else:
            vals.append(str(c.value).strip())

    # Trim trailing empty columns
    while vals and vals[-1] == "":
        vals.pop()

    return f"Row {row_idx:3d}: {tag_str}| {' | '.join(vals)} |"
