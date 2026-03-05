"""
Annotate the client's funds_flow.xlsx with match-status columns.

Opens the original Excel, adds audit annotation columns to the right of each
in-scope tab, and returns the workbook (unsaved) along with row-level info
needed by the Journal Entry tab builder.
"""
from __future__ import annotations

import json
import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

from agent.output.styles import (
    BOLD_FONT, CELL_FONT, HEADER_FILL, HEADER_FONT,
    STATUS_COLORS, THIN_BORDER, USD_FMT, CHK_FMT,
)

_COA_PATH = Path(__file__).parent.parent.parent / "chart_of_accounts.json"

_DEFAULT_CREDIT = {
    "wire_payable":        ("2100", "Wire Payable"),
    "accrued_liabilities": ("2200", "Accrued Liabilities — Deal Costs"),
}


def load_credit_accounts() -> dict:
    """Load credit-side account codes from chart_of_accounts.json."""
    if not _COA_PATH.exists():
        return _DEFAULT_CREDIT
    try:
        coa = json.loads(_COA_PATH.read_text())
        credits = {
            k: (v["code"], v["name"])
            for k, v in coa.get("credit_accounts", {}).items()
        }
        return credits or _DEFAULT_CREDIT
    except Exception as e:
        print(f"  Warning: could not load chart_of_accounts.json ({e}) — using defaults")
        return _DEFAULT_CREDIT


def _get_gl(item: dict) -> tuple[str, str]:
    code = item.get("gl_account_code") or "7120"
    name = item.get("gl_account_name") or "Transaction Costs — Miscellaneous Closing Costs"
    return code, name


def _detect_total_col(ws) -> int | None:
    """Return 1-based column number of the Total/Amount column, or None."""
    for row in ws.iter_rows(max_row=6, values_only=True):
        for i, cell in enumerate(row, start=1):
            if cell and "total" in str(cell).lower():
                return i
    return None


def annotate(deal_dir: Path, index: dict) -> tuple[openpyxl.Workbook, Path, list[dict]]:
    """
    Copy funds_flow.xlsx -> funds_flow_indexed.xlsx, add annotation columns.

    Returns (workbook, output_path, je_row_info).
    Caller is responsible for adding further sheets and saving.
    """
    ff_path  = deal_dir / "funds_flow.xlsx"
    out_path = deal_dir / "run_output" / "funds_flow_indexed.xlsx"

    shutil.copy(str(ff_path), str(out_path))
    wb = openpyxl.load_workbook(str(out_path))

    lookup = {
        item["description"].strip().lower(): item
        for item in index["line_items"]
    }

    fi_pct  = index["fund_allocations"].get("Fund I",  0.55)
    fii_pct = index["fund_allocations"].get("Fund II", 0.45)

    je_row_info: list[dict] = []
    desc_to_je: dict[str, dict] = {}

    for item in index["line_items"]:
        key = item["description"].strip().lower()
        acct_code, acct_name = _get_gl(item)
        rec = {
            "ff_ref":      item.get("ff_ref"),
            "status":      item.get("status", "MISSING"),
            "acct_code":   acct_code,
            "acct_name":   acct_name,
            "description": item["description"],
            "sheet":       None,
            "row":         None,
            "fi_col":      None,
            "fii_col":     None,
            "ff_amount":   item["funds_flow_amount"],
            "doc_amount":  item.get("document_amount") or 0.0,
            "vendor":      item.get("document_vendor") or item["description"],
        }
        je_row_info.append(rec)
        desc_to_je[key] = rec

    for ws in wb.worksheets:
        title_lower = ws.title.lower()
        if any(k in title_lower for k in ("source", "wire", "instruction")):
            continue

        total_col_1 = _detect_total_col(ws)
        max_col = ws.max_column or 5
        ann = max_col + 2

        headers = [
            "FF Ref", "Match Status", "Document", "Amount Match", "Notes",
            "GL Account", f"Fund I ({fi_pct*100:.0f}%)", f"Fund II ({fii_pct*100:.0f}%)", "Σ Check",
        ]
        col_offsets = [0, 1, 2, 3, 4, 6, 7, 8, 9]

        for offset, h in zip(col_offsets, headers):
            c = ws.cell(1, ann + offset, h)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        col_widths = {0: 7, 1: 12, 2: 40, 3: 12, 4: 45, 6: 36, 7: 14, 8: 14, 9: 12}
        for offset, w in col_widths.items():
            ws.column_dimensions[get_column_letter(ann + offset)].width = w

        for xl_row in ws.iter_rows(min_row=2):
            desc_cell = xl_row[0]
            if not desc_cell.value:
                continue
            key = str(desc_cell.value).strip().lower()
            if key not in lookup:
                continue

            item   = lookup[key]
            status = (item.get("status") or "").upper()
            color  = STATUS_COLORS.get(status, "FFFFFF")
            fill   = PatternFill("solid", fgColor=color)
            r      = desc_cell.row

            ff_ref_val = item.get("ff_ref") or "—"

            std_vals = [
                (ann,     ff_ref_val),
                (ann + 1, status),
                (ann + 2, item.get("document_file") or ""),
                (ann + 3, "YES" if item.get("amount_agrees") is True
                          else "NO" if item.get("amount_agrees") is False
                          else "N/A"),
                (ann + 4, item.get("notes") or ""),
            ]
            for col, val in std_vals:
                c = ws.cell(r, col, val)
                c.fill = fill
                c.font = CELL_FONT
                c.alignment = Alignment(vertical="center", wrap_text=(col == ann + 4))

            acct_code, acct_name = _get_gl(item)
            gl_cell = ws.cell(r, ann + 6, f"{acct_code}  {acct_name}")
            gl_cell.fill = fill
            gl_cell.font = CELL_FONT
            gl_cell.alignment = Alignment(vertical="center")

            if total_col_1:
                total_letter = get_column_letter(total_col_1)
                fi_letter    = get_column_letter(ann + 7)
                fii_letter   = get_column_letter(ann + 8)
                chk_letter   = get_column_letter(ann + 9)

                fi_cell = ws.cell(r, ann + 7, f"={total_letter}{r}*{fi_pct}")
                fi_cell.fill = fill
                fi_cell.font = CELL_FONT
                fi_cell.number_format = USD_FMT
                fi_cell.alignment = Alignment(horizontal="right", vertical="center")

                fii_cell = ws.cell(r, ann + 8, f"={total_letter}{r}*{fii_pct}")
                fii_cell.fill = fill
                fii_cell.font = CELL_FONT
                fii_cell.number_format = USD_FMT
                fii_cell.alignment = Alignment(horizontal="right", vertical="center")

                chk_cell = ws.cell(r, ann + 9,
                                   f"={fi_letter}{r}+{fii_letter}{r}-{total_letter}{r}")
                chk_cell.fill = fill
                chk_cell.font = CELL_FONT
                chk_cell.number_format = CHK_FMT
                chk_cell.alignment = Alignment(horizontal="right", vertical="center")

                if key in desc_to_je:
                    rec = desc_to_je[key]
                    rec["sheet"]   = ws.title
                    rec["row"]     = r
                    rec["fi_col"]  = fi_letter
                    rec["fii_col"] = fii_letter

    return wb, out_path, je_row_info
