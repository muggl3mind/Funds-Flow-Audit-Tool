"""
Stage 6b: Annotate the client's funds_flow.xlsx in-place.

Auditors do not create a separate file — they add their tick marks and
support references directly to the client's workbook.

What this writer does:
  1. Opens the client's original funds_flow.xlsx (preserving all formulas and formatting)
  2. For each in-scope tab that has matched line items, adds audit columns to the RIGHT
     of the existing data:
       [Support Document] [Doc Amount] [Agrees?] [Confidence] [Flags] [Auditor Note]
  3. Adds an "Audit Summary" sheet at the end of the workbook
  4. Saves as funds_flow_indexed.xlsx (original is never overwritten)
"""
from __future__ import annotations

import shutil
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from agent.config import DealConfig
from agent.exceptions.exception_classifier import DealException
from agent.matcher.llm_matcher import MatchResult
from agent.normalizers.document_normalizer import DocumentRecord
from agent.normalizers.funds_flow_normalizer import TabScopeResult
from agent.utils.amount_utils import format_usd

# ── Colours ──────────────────────────────────────────────────────────────────
C_NAVY       = "1F3864"
C_WHITE      = "FFFFFF"
C_GREEN_BG   = "E2EFDA"
C_GREEN_FG   = "375623"
C_RED_BG     = "FCE4D6"
C_RED_FG     = "C00000"
C_AMBER_BG   = "FFF2CC"
C_AMBER_FG   = "7F6000"
C_BLUE_BG    = "D6E4F0"
C_GREY_BG    = "F2F2F2"
C_PURPLE_BG  = "EAD1DC"   # superseded invoices
C_AUDIT_HDR  = "2E4A7A"   # audit column header stripe

SEV_COLORS = {"HIGH": C_RED_BG, "MEDIUM": C_AMBER_BG, "LOW": C_GREY_BG, "INFO": C_BLUE_BG}

AUDIT_COLS = [
    ("Support Document",  26),
    ("Doc Amount",        14),
    ("Agrees?",           10),
    ("Confidence",        11),
    ("Exception Flags",   28),
    ("Auditor Note",      42),
]

STATUS_SYMBOL = {
    "matched":   "✓",
    "partial":   "~",
    "exception": "!",
    "missing":   "✗",
}


def _side(style="thin"):
    return Side(style=style)

def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=10, color="000000", italic=False) -> Font:
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def _align(h="left", v="center", wrap=False, indent=0) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)


# ── Public entry point ────────────────────────────────────────────────────────

def annotate_client_file(
    source_path: Path,
    results: list[MatchResult],
    exceptions: list[DealException],
    scope: TabScopeResult,
    all_documents: list[DocumentRecord],
    config: DealConfig,
    summary: dict,
    output_path: Path,
):
    """
    Copy source_path → output_path, then annotate with audit columns in-place.
    """
    shutil.copy2(source_path, output_path)

    # Open WITHOUT data_only so we preserve formulas
    wb = openpyxl.load_workbook(str(output_path))

    # Group results by source tab
    by_tab: dict[str, list[MatchResult]] = {}
    for r in results:
        by_tab.setdefault(r.line_item.source_tab, []).append(r)

    for tab_name, tab_results in by_tab.items():
        if tab_name not in wb.sheetnames:
            continue
        ws = wb[tab_name]
        _annotate_tab(ws, tab_results)

    _add_audit_summary(wb, results, exceptions, scope, all_documents, config, summary)

    wb.save(str(output_path))


# ── Tab annotation ────────────────────────────────────────────────────────────

def _annotate_tab(ws, results: list[MatchResult]):
    """Add audit columns to the right of the existing data in one worksheet."""

    # Find the rightmost used column so we know where to start audit columns
    max_col = ws.max_column or 1
    audit_start_col = max_col + 2   # leave one blank column as visual separator

    # Blank separator column
    sep_col = max_col + 1
    ws.column_dimensions[get_column_letter(sep_col)].width = 2

    # Write audit column headers
    # Find the header row: look for any row in the first 10 that has bold/header-like cells
    # Heuristic: use row 3 if it exists with content, else row 1
    header_row = _find_header_row(ws)

    # Merge a label above the audit columns
    label_row = max(1, header_row - 1)
    merge_end_col = audit_start_col + len(AUDIT_COLS) - 1
    try:
        ws.merge_cells(
            start_row=label_row, start_column=audit_start_col,
            end_row=label_row, end_column=merge_end_col
        )
    except Exception:
        pass
    lc = ws.cell(label_row, audit_start_col)
    lc.value = "AUDIT — SUPPORT DOCUMENT INDEX"
    lc.font = _font(bold=True, color=C_WHITE, size=10)
    lc.fill = _fill(C_AUDIT_HDR)
    lc.alignment = _align(h="center")
    lc.border = _border()

    for col_offset, (header, width) in enumerate(AUDIT_COLS):
        col = audit_start_col + col_offset
        cell = ws.cell(header_row, col, header)
        cell.font = _font(bold=True, color=C_WHITE)
        cell.fill = _fill(C_AUDIT_HDR)
        cell.alignment = _align(h="center", wrap=True)
        cell.border = _border()
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[header_row].height = max(ws.row_dimensions[header_row].height or 0, 28)

    # Build a map: row_idx → MatchResult
    row_to_result: dict[int, MatchResult] = {}
    for r in results:
        row_idx = _parse_row_idx(r.line_item.source_row)
        if row_idx is not None:
            row_to_result[row_idx] = r

    # Write audit values for each matched row
    for row_idx, r in row_to_result.items():
        _write_audit_row(ws, row_idx, r, audit_start_col)

    # Add a "superseded" note for any K&E-style superseded docs visible in the sheet
    # (these appear in the exceptions list; mark any rows not in results as reviewed)


def _write_audit_row(ws, row_idx: int, r: MatchResult, audit_start_col: int):
    doc  = r.matched_document
    item = r.line_item

    # Background colour for the audit cells matches status
    bg = (C_GREEN_BG  if r.status == "matched"   else
          C_AMBER_BG  if r.status in ("partial", "exception") else
          C_RED_BG    if r.status == "missing"   else
          C_PURPLE_BG)   # superseded

    symbol = STATUS_SYMBOL.get(r.status, "?")
    agrees_str = {True: "✓ Yes", False: "✗ No", None: "—"}[r.amount_agrees]
    conf_str   = f"{r.confidence_score:.0%}" if r.confidence_score else "—"
    flags_str  = ", ".join(r.exception_flags[:3]) if r.exception_flags else ""
    doc_file   = doc.file_name if doc else "— NO DOCUMENT —"
    doc_amt    = doc.total_amount if doc else None

    row_data = [
        f"{symbol}  {doc_file}",
        doc_amt,
        agrees_str,
        conf_str,
        flags_str,
        r.notes[:160] if r.notes else "",
    ]

    for col_offset, value in enumerate(row_data):
        col = audit_start_col + col_offset
        cell = ws.cell(row_idx, col, value)
        cell.fill = _fill(bg)
        cell.border = _border()
        cell.font = _font(
            bold=(col_offset == 0),
            color=(C_GREEN_FG if r.status == "matched" else
                   C_RED_FG   if r.status == "missing" else "000000"),
        )
        cell.alignment = _align(
            h="right" if col_offset == 1 else "center" if col_offset in (2, 3) else "left",
            wrap=(col_offset == 5),
        )
        if col_offset == 1 and value is not None:
            cell.number_format = '"$"#,##0'

    # Also highlight the original description cell(s) with a thin left border in the status colour
    # to create a visual tie-mark on the same row
    _add_tick_mark(ws, row_idx, r.status)


def _add_tick_mark(ws, row_idx: int, status: str):
    """Add a coloured left border on cell A of the line item row as a visual tick mark."""
    color = (C_GREEN_FG if status == "matched"  else
             C_AMBER_FG if status in ("partial", "exception") else
             C_RED_FG)
    cell = ws.cell(row_idx, 1)
    existing = cell.border
    new_border = Border(
        left=Side(style="medium", color=color),
        right=existing.right if existing else _side(),
        top=existing.top    if existing else _side(),
        bottom=existing.bottom if existing else _side(),
    )
    cell.border = new_border


def _find_header_row(ws) -> int:
    """Heuristic: find the first row that has 3+ non-blank cells — likely the header."""
    for row_idx in range(1, min(8, (ws.max_row or 8) + 1)):
        non_blank = sum(
            1 for col in range(1, (ws.max_column or 1) + 1)
            if ws.cell(row_idx, col).value is not None
        )
        if non_blank >= 3:
            return row_idx
    return 3


def _parse_row_idx(source_row: str) -> Optional[int]:
    """Extract integer row number from 'Buyer Expenses:Row5' or 'Row5'."""
    try:
        part = source_row.split(":")[-1]
        return int(part.replace("Row", "").replace("row", "").strip())
    except (ValueError, AttributeError):
        return None


# ── Audit Summary sheet ───────────────────────────────────────────────────────

def _add_audit_summary(
    wb,
    results: list[MatchResult],
    exceptions: list[DealException],
    scope: TabScopeResult,
    all_documents: list[DocumentRecord],
    config: DealConfig,
    summary: dict,
):
    # Remove existing audit summary if re-running
    if "Audit Summary" in wb.sheetnames:
        del wb["Audit Summary"]

    ws = wb.create_sheet("Audit Summary")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 40

    # ── Title ──
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"{config.deal_name} — Audit Support Index"
    c.font = _font(bold=True, size=13, color=C_WHITE)
    c.fill = _fill(C_NAVY)
    c.alignment = _align(h="center")
    ws.row_dimensions[1].height = 24

    # ── Deal metadata ──
    ws.merge_cells("A2:G2")
    ws["A2"].value = (
        f"Closing: {config.closing_date}  |  "
        f"Indexed: {summary.get('indexed_at', '')}  |  "
        f"Client role: {config.client_role.upper()}  |  "
        f"Agent v{config.agent_version}"
    )
    ws["A2"].font = _font(italic=True, size=9, color=C_WHITE)
    ws["A2"].fill = _fill(C_AUDIT_HDR)
    ws["A2"].alignment = _align(h="center")
    ws.row_dimensions[2].height = 14

    ws.row_dimensions[3].height = 6

    # ── Scoreboard ──
    score_items = [
        ("Total Line Items",        summary["total_line_items"],  None),
        ("Matched",                 summary["matched"],           C_GREEN_FG),
        ("Partial",                 summary["partial"],           C_AMBER_FG),
        ("Missing",                 summary["missing"],           C_RED_FG),
        ("Exceptions",              summary.get("exceptions", 0), C_AMBER_FG),
        ("Amount Mismatches",       summary["amount_mismatches"], C_RED_FG if summary["amount_mismatches"] else C_GREEN_FG),
        ("Total Supported ($)",     summary["total_supported_amount"],   None),
        ("Total Unsupported ($)",   summary["total_unsupported_amount"], C_RED_FG),
        ("Avg Match Confidence",    summary.get("match_confidence_avg"), None),
    ]
    for r_off, (label, value, fg) in enumerate(score_items, start=4):
        lc = ws.cell(r_off, 1, label)
        vc = ws.cell(r_off, 2, value)
        lc.font = _font(bold=True)
        vc.font = _font(bold=True, color=fg or "000000")
        for c in (lc, vc):
            c.fill = _fill(C_BLUE_BG)
            c.border = _border()
            c.alignment = _align(indent=1)
        if label in ("Total Supported ($)", "Total Unsupported ($)") and value:
            vc.number_format = '"$"#,##0'
        if label == "Avg Match Confidence" and value:
            vc.number_format = "0%"
        ws.row_dimensions[r_off].height = 16

    # ── Tabs in scope ──
    r = len(score_items) + 5
    ws.cell(r, 1, "Tabs In Scope").font = _font(bold=True)
    ws.cell(r, 2, ", ".join(scope.in_scope) or "(none)").font = _font(color=C_GREEN_FG)
    r += 1
    ws.cell(r, 1, "Tabs Skipped").font = _font(bold=True)
    ws.cell(r, 2, ", ".join(scope.skipped) or "(none)").font = _font(color=C_RED_FG)
    for row in (r - 1, r):
        ws.row_dimensions[row].height = 14

    r += 2

    # ── Full index table ──
    ws.row_dimensions[r].height = 6
    r += 1

    headers = ["Description", "Fund Allocations", "FF Amount",
               "Doc Amount", "Agrees?", "Status", "Document / Note"]
    hdrs_w  = [32, 20, 14, 14, 10, 12, 46]
    for col, (h, w) in enumerate(zip(headers, hdrs_w), 1):
        cell = ws.cell(r, col, h)
        cell.font = _font(bold=True, color=C_WHITE)
        cell.fill = _fill(C_NAVY)
        cell.alignment = _align(h="center")
        cell.border = _border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[r].height = 20
    r += 1

    for result in results:
        item = result.line_item
        doc  = result.matched_document
        bg   = (C_GREEN_BG  if result.status == "matched"  else
                C_AMBER_BG  if result.status in ("partial", "exception") else
                C_RED_BG)

        alloc_str = "  |  ".join(
            f"{k}: {format_usd(v)}" for k, v in item.fund_allocations.items()
        ) if item.fund_allocations else "—"
        agrees_str = {True: "✓ Yes", False: "✗ No", None: "—"}[result.amount_agrees]
        status_str = f"{STATUS_SYMBOL.get(result.status,'?')}  {result.status.upper()}"
        doc_str    = doc.file_name if doc else "— NO DOCUMENT —"

        row_data = [
            item.description, alloc_str,
            item.total_amount, doc.total_amount if doc else None,
            agrees_str, status_str, doc_str,
        ]
        status_fg = (C_GREEN_FG if result.status == "matched" else
                     C_RED_FG   if result.status == "missing" else "000000")

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(r, col, value)
            cell.fill = _fill(bg)
            cell.border = _border()
            cell.font = _font(
                bold=(col == 6),
                color=status_fg if col == 6 else "000000",
            )
            cell.alignment = _align(
                h="right"  if col in (3, 4) else
                "center" if col in (5, 6) else "left"
            )
            if col in (3, 4) and value is not None:
                cell.number_format = '"$"#,##0'
        ws.row_dimensions[r].height = 16
        r += 1

    # ── Exceptions ──
    r += 1
    ws.merge_cells(f"A{r}:G{r}")
    ws.cell(r, 1, "EXCEPTIONS & ACTION ITEMS").font = _font(bold=True, color=C_WHITE, size=11)
    ws.cell(r, 1).fill = _fill(C_NAVY)
    ws.cell(r, 1).alignment = _align(h="left", indent=1)
    ws.row_dimensions[r].height = 18
    r += 1

    exc_headers = ["Severity", "Type", "Source", "Description", "Suggested Action"]
    exc_widths  = [10, 24, 26, 44, 44]
    for col, (h, w) in enumerate(zip(exc_headers, exc_widths), 1):
        cell = ws.cell(r, col, h)
        cell.font = _font(bold=True, color=C_WHITE)
        cell.fill = _fill(C_AUDIT_HDR)
        cell.border = _border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[r].height = 18
    r += 1

    if not exceptions:
        ws.cell(r, 1, "No exceptions found.").font = _font(italic=True)
    else:
        for exc in exceptions:
            bg = SEV_COLORS.get(exc.severity, C_GREY_BG)
            for col, val in enumerate(
                [exc.severity, exc.exception_type, exc.source_ref,
                 exc.description, exc.suggested_action], 1
            ):
                cell = ws.cell(r, col, val)
                cell.fill = _fill(bg)
                cell.border = _border()
                cell.font = _font(bold=(col == 1))
                cell.alignment = _align(wrap=(col in (4, 5)), h="center" if col <= 2 else "left")
            ws.row_dimensions[r].height = 32
            r += 1

    ws.freeze_panes = "A4"
