"""
agent/write_journal_entry.py — Export the journal entry to Excel.

Creates <deal_dir>/run_output/journal_entry.xlsx with two sheets:
  - Journal Entry    — full debit/credit table
  - Accrual Schedule — MISSING/PARTIAL items with follow-up status

Usage:
  python agent/write_journal_entry.py <deal_dir>
"""
from __future__ import annotations

import json
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Styles ────────────────────────────────────────────────────────────────────

NAVY        = "1F3864"
LIGHT_BLUE  = "DDEBF7"
LIGHT_GREEN = "C6EFCE"
LIGHT_AMBER = "FFEB9C"
LIGHT_RED   = "FFC7CE"
LIGHT_GREY  = "F2F2F2"
WHITE       = "FFFFFF"

HEADER_FILL  = PatternFill("solid", fgColor=NAVY)
HEADER_FONT  = Font(name="Calibri", bold=True, color=WHITE, size=10)
TITLE_FONT   = Font(name="Calibri", bold=True, size=12)
BOLD_FONT    = Font(name="Calibri", bold=True, size=10)
CELL_FONT    = Font(name="Calibri", size=10)
ITALIC_FONT  = Font(name="Calibri", italic=True, size=10)

THIN_SIDE    = Side(style="thin", color="BFBFBF")
THIN_BORDER  = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
BOTTOM_BORDER = Border(bottom=Side(style="medium", color=NAVY))

STATUS_FILLS = {
    "MATCHED":    PatternFill("solid", fgColor=LIGHT_GREEN),
    "CUMULATIVE": PatternFill("solid", fgColor=LIGHT_BLUE),
    "PARTIAL":    PatternFill("solid", fgColor=LIGHT_AMBER),
    "MISSING":    PatternFill("solid", fgColor=LIGHT_RED),
}


def _set_col_width(ws, col: int, width: float) -> None:
    ws.column_dimensions[get_column_letter(col)].width = width


def _header_row(ws, row: int, values: list[str], widths: list[float] | None = None) -> None:
    for c, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill  = HEADER_FILL
        cell.font  = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    if widths:
        for c, w in enumerate(widths, start=1):
            _set_col_width(ws, c, w)


def _data_cell(ws, row: int, col: int, value, *, bold=False, align="left",
               fill: PatternFill | None = None, number_format: str | None = None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font   = BOLD_FONT if bold else CELL_FONT
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill
    if number_format:
        cell.number_format = number_format


# ── Account mapping ───────────────────────────────────────────────────────────

ACCOUNT_RULES: list[tuple[list[str], str]] = [
    (["Legal", "Counsel", "SPA"],                     "Transaction Costs — Legal Fees"),
    (["Banking", "Advisory", "M&A", "Fairness"],      "Transaction Costs — Banking & Advisory"),
    (["Quality of Earnings", "QofE", "Financial DD"], "Transaction Costs — Financial Due Diligence"),
    (["IT", "Cybersecurity", "Technology", "Systems"],"Transaction Costs — Technology Due Diligence"),
    (["Environmental", "ESG"],                        "Transaction Costs — Environmental Due Diligence"),
    (["Operational", "Commercial", "Strategy"],       "Transaction Costs — Operational Due Diligence"),
    (["Tax", "Structuring"],                          "Transaction Costs — Tax Advisory"),
    (["HSR", "Regulatory", "Filing"],                 "Transaction Costs — Regulatory Fees"),
    (["Travel", "T&E", "Expenses"],                   "Transaction Costs — Travel & Expenses"),
    (["Miscellaneous", "Closing Costs", "Notary"],    "Transaction Costs — Miscellaneous"),
]


def _map_account(description: str) -> str:
    desc_lower = description.lower()
    for keywords, account in ACCOUNT_RULES:
        if any(kw.lower() in desc_lower for kw in keywords):
            return account
    return "Transaction Costs — Miscellaneous"


# ── Wire Payable calculation for PARTIAL items ────────────────────────────────

def _partial_wire_amounts(item: dict) -> tuple[float, float]:
    """Return (fund_i_wire, fund_ii_wire) for a PARTIAL item based on doc amount."""
    doc_amt  = item.get("document_amount") or 0.0
    total    = item["funds_flow_amount"]
    fi_ratio = item["fund_i_amount"] / total if total else 0.55
    fii_ratio = 1.0 - fi_ratio
    return round(doc_amt * fi_ratio, 2), round(doc_amt * fii_ratio, 2)


# ── Sheet 1: Journal Entry ────────────────────────────────────────────────────

def _build_je_sheet(ws, index: dict) -> None:
    deal        = index["deal"]
    closing_dt  = index["closing_date"]
    prepared_dt = str(date.today())
    line_items  = index["line_items"]

    # ── Title block ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value = f"JOURNAL ENTRY — Deal Transaction Costs   |   {deal}"
    title.font  = TITLE_FONT
    title.fill  = HEADER_FILL
    title.font  = Font(name="Calibri", bold=True, color=WHITE, size=12)
    title.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 22

    meta = [
        ("Target:", "NovaTech Solutions"),
        ("Period:", closing_dt),
        ("Ref:",    "FF-JE-001"),
        ("Prepared:", prepared_dt),
        ("Status:", "⚠  DRAFT — For review by fund accounting before posting"),
    ]
    for i, (label, val) in enumerate(meta, start=2):
        ws.cell(row=i, column=1, value=label).font = BOLD_FONT
        cell = ws.cell(row=i, column=2, value=val)
        cell.font = CELL_FONT
        if label == "Status:":
            cell.font = Font(name="Calibri", bold=True, color="C00000", size=10)
    ws.merge_cells("B6:G6")  # status spans

    # ── DEBIT section ─────────────────────────────────────────────────────────
    DEBIT_ROW = 8
    ws.cell(row=DEBIT_ROW - 1, column=1, value="DEBIT LINES").font = BOLD_FONT
    _header_row(ws, DEBIT_ROW,
                ["FF Ref", "Account", "Fund", "Amount", "Vendor / Description", "Status", "Notes"],
                widths=[8, 42, 9, 14, 32, 12, 30])
    ws.row_dimensions[DEBIT_ROW].height = 18

    row = DEBIT_ROW + 1
    for item in line_items:
        account = _map_account(item["description"])
        status  = item["status"]
        ff_ref  = item.get("ff_ref") or "—"
        vendor  = item.get("document_vendor") or item["description"]
        fill    = STATUS_FILLS.get(status)

        for fund, amount in [("Fund I",  item["fund_i_amount"]),
                              ("Fund II", item["fund_ii_amount"])]:
            _data_cell(ws, row, 1, ff_ref,   align="center", fill=fill)
            _data_cell(ws, row, 2, account,  fill=fill)
            _data_cell(ws, row, 3, fund,     align="center", fill=fill)
            _data_cell(ws, row, 4, amount,   align="right",  fill=fill, number_format='"$"#,##0.00')
            _data_cell(ws, row, 5, vendor,   fill=fill)
            _data_cell(ws, row, 6, status,   align="center", fill=fill)
            _data_cell(ws, row, 7, item.get("notes", "") if fund == "Fund I" else "", fill=fill)
            row += 1

    # Total debits
    total_debits = sum(i["funds_flow_amount"] for i in line_items)
    ws.cell(row=row, column=3, value="Total Debits").font = BOLD_FONT
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="right")
    cell = ws.cell(row=row, column=4, value=total_debits)
    cell.font = BOLD_FONT
    cell.alignment = Alignment(horizontal="right")
    cell.number_format = '"$"#,##0.00'
    cell.border = BOTTOM_BORDER
    row += 2

    # ── CREDIT section ────────────────────────────────────────────────────────
    ws.cell(row=row - 1, column=1, value="CREDIT LINES").font = BOLD_FONT
    _header_row(ws, row,
                ["Account", "Fund", "Amount", "Basis", "Items Included"])
    ws.row_dimensions[row].height = 18
    # override widths for credit table (same columns A-E)
    row += 1

    # Accumulate wire payable and accrued liabilities by fund
    wp_fi = wp_fii = 0.0
    ac_fi = ac_fii = 0.0

    wp_items: list[str] = []
    ac_items: list[str] = []

    for item in line_items:
        status = item["status"]
        ff_ref = item.get("ff_ref") or "—"
        desc   = item["description"]
        label  = f"{ff_ref} {desc}" if ff_ref != "—" else desc

        if status in ("MATCHED", "CUMULATIVE"):
            wp_fi  += item["fund_i_amount"]
            wp_fii += item["fund_ii_amount"]
            wp_items.append(f"{ff_ref} — {desc}")

        elif status == "PARTIAL":
            fi_wire, fii_wire = _partial_wire_amounts(item)
            fi_acc  = item["fund_i_amount"]  - fi_wire
            fii_acc = item["fund_ii_amount"] - fii_wire
            wp_fi  += fi_wire
            wp_fii += fii_wire
            ac_fi  += fi_acc
            ac_fii += fii_acc
            wp_items.append(f"{ff_ref} — {desc} (partial ${item['document_amount']:,.0f})")
            ac_items.append(f"{ff_ref} — {desc} (shortfall ${item['funds_flow_amount'] - item['document_amount']:,.0f})")

        elif status == "MISSING":
            ac_fi  += item["fund_i_amount"]
            ac_fii += item["fund_ii_amount"]
            ac_items.append(f"— {desc}")

    wp_basis = "Matched/Cumulative invoices on file"
    ac_basis = "Missing invoices / partial shortfalls — see Accrual Schedule"

    credit_rows = [
        ("Wire Payable",        "Fund I",  wp_fi,  wp_basis, "; ".join(wp_items[:3]) + ("…" if len(wp_items) > 3 else "")),
        ("Wire Payable",        "Fund II", wp_fii, wp_basis, ""),
        ("Accrued Liabilities", "Fund I",  ac_fi,  ac_basis, "; ".join(ac_items)),
        ("Accrued Liabilities", "Fund II", ac_fii, ac_basis, ""),
    ]

    for acct, fund, amount, basis, items_str in credit_rows:
        fill = PatternFill("solid", fgColor=LIGHT_GREEN) if "Wire" in acct else PatternFill("solid", fgColor=LIGHT_AMBER)
        _data_cell(ws, row, 1, acct,      bold=True, fill=fill)
        _data_cell(ws, row, 2, fund,      align="center", fill=fill)
        _data_cell(ws, row, 3, amount,    align="right",  fill=fill, number_format='"$"#,##0.00')
        _data_cell(ws, row, 4, basis,     fill=fill)
        _data_cell(ws, row, 5, items_str, fill=fill)
        row += 1

    # Total credits
    total_credits = wp_fi + wp_fii + ac_fi + ac_fii
    ws.cell(row=row, column=2, value="Total Credits").font = BOLD_FONT
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
    cell = ws.cell(row=row, column=3, value=total_credits)
    cell.font = BOLD_FONT
    cell.alignment = Alignment(horizontal="right")
    cell.number_format = '"$"#,##0.00'
    cell.border = BOTTOM_BORDER
    row += 2

    # ── Balance check ─────────────────────────────────────────────────────────
    balanced = abs(total_debits - total_credits) < 0.01
    check_val = f"✓  BALANCED  (${total_debits:,.2f})" if balanced else f"✗  OUT OF BALANCE  Dr ${total_debits:,.2f}  Cr ${total_credits:,.2f}"
    check_fill = PatternFill("solid", fgColor=LIGHT_GREEN) if balanced else PatternFill("solid", fgColor=LIGHT_RED)
    ws.merge_cells(f"A{row}:G{row}")
    cell = ws.cell(row=row, column=1, value=check_val)
    cell.font  = Font(name="Calibri", bold=True, size=11)
    cell.fill  = check_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 20
    row += 2

    # ── Fund summary ──────────────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="FUND SUMMARY").font = BOLD_FONT
    row += 1
    _header_row(ws, row, ["Fund", "Total Costs", "Wire Payable", "Accrued Liabilities"])
    row += 1

    fund_i_total  = sum(i["fund_i_amount"]  for i in line_items)
    fund_ii_total = sum(i["fund_ii_amount"] for i in line_items)

    for label, total, wp, ac in [
        ("Fund I",  fund_i_total,  wp_fi,  ac_fi),
        ("Fund II", fund_ii_total, wp_fii, ac_fii),
        ("Total",   fund_i_total + fund_ii_total, wp_fi + wp_fii, ac_fi + ac_fii),
    ]:
        bold = (label == "Total")
        fill = PatternFill("solid", fgColor=LIGHT_GREY) if bold else None
        _data_cell(ws, row, 1, label,  bold=bold, align="center", fill=fill)
        _data_cell(ws, row, 2, total,  bold=bold, align="right",  fill=fill, number_format='"$"#,##0.00')
        _data_cell(ws, row, 3, wp,     bold=bold, align="right",  fill=fill, number_format='"$"#,##0.00')
        _data_cell(ws, row, 4, ac,     bold=bold, align="right",  fill=fill, number_format='"$"#,##0.00')
        row += 1

    ws.freeze_panes = "A9"


# ── Sheet 2: Accrual Schedule ─────────────────────────────────────────────────

def _build_accrual_sheet(ws, index: dict) -> None:
    deal       = index["deal"]
    line_items = index["line_items"]

    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value = f"ACCRUAL SCHEDULE — Items Requiring Follow-Up   |   {deal}"
    title.font  = Font(name="Calibri", bold=True, color=WHITE, size=12)
    title.fill  = HEADER_FILL
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.cell(row=2, column=1, value=(
        "These items were not fully supported at closing. "
        "Do not post Wire Payable until invoices are received. "
        "Reverse accrual when matched."
    )).font = ITALIC_FONT
    ws.merge_cells("A2:G2")

    _header_row(ws, 4,
                ["FF Ref", "Vendor", "Description", "FF Amount",
                 "Supported", "Shortfall", "Action Required"],
                widths=[8, 28, 36, 14, 14, 14, 38])
    ws.row_dimensions[4].height = 18

    row = 5
    total_accrual = 0.0

    for item in line_items:
        if item["status"] not in ("MISSING", "PARTIAL"):
            continue

        ff_ref    = item.get("ff_ref") or "—"
        vendor    = item.get("document_vendor") or item.get("notes", "Unknown")
        desc      = item["description"]
        ff_amt    = item["funds_flow_amount"]
        doc_amt   = item.get("document_amount") or 0.0
        shortfall = ff_amt - doc_amt
        total_accrual += shortfall

        if item["status"] == "MISSING":
            action = "Obtain invoice from vendor. Post Wire Payable on receipt. Reverse accrual."
        else:
            action = f"SEC Form D receipt outstanding. ${shortfall:,.0f} not yet supported."

        fill = STATUS_FILLS.get(item["status"])
        _data_cell(ws, row, 1, ff_ref,   align="center", fill=fill)
        _data_cell(ws, row, 2, vendor,   fill=fill)
        _data_cell(ws, row, 3, desc,     fill=fill)
        _data_cell(ws, row, 4, ff_amt,   align="right", fill=fill, number_format='"$"#,##0.00')
        _data_cell(ws, row, 5, doc_amt,  align="right", fill=fill, number_format='"$"#,##0.00')
        _data_cell(ws, row, 6, shortfall,align="right", fill=fill, number_format='"$"#,##0.00')
        _data_cell(ws, row, 7, action,   fill=fill)
        ws.row_dimensions[row].height = 16
        row += 1

    row += 1
    ws.cell(row=row, column=5, value="Total Accrual").font = BOLD_FONT
    ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")
    cell = ws.cell(row=row, column=6, value=total_accrual)
    cell.font = BOLD_FONT
    cell.alignment = Alignment(horizontal="right")
    cell.number_format = '"$"#,##0.00'
    cell.border = BOTTOM_BORDER

    row += 2
    note = ws.cell(row=row, column=1, value=(
        "Note: This entry covers transaction costs only. "
        "The investment in target (purchase price) is a separate entry "
        "booked from the wire confirmation / SPA."
    ))
    note.font = ITALIC_FONT
    ws.merge_cells(f"A{row}:G{row}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: python agent/write_journal_entry.py <deal_dir>")
        sys.exit(1)

    deal_dir   = Path(sys.argv[1])
    index_path = deal_dir / "run_output" / "index.json"

    if not index_path.exists():
        print(f"ERROR: {index_path} not found. Run /index-funds-flow first.")
        sys.exit(1)

    index = json.loads(index_path.read_text())

    wb = Workbook()

    # Sheet 1
    ws_je = wb.active
    ws_je.title = "Journal Entry"
    _build_je_sheet(ws_je, index)

    # Sheet 2
    ws_ac = wb.create_sheet("Accrual Schedule")
    _build_accrual_sheet(ws_ac, index)

    out_path = deal_dir / "run_output" / "journal_entry.xlsx"
    wb.save(str(out_path))
    print(f"  Saved: {out_path}")


if __name__ == "__main__":
    main()
