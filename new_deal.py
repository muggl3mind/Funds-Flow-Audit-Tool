"""
new_deal.py — Scaffold a new deal folder for the Funds Flow Audit Tool.

Creates the following structure:
  deals/<deal_slug>/
    documents/           ← drop support PDFs here before running the agent
    run_output/          ← agent writes outputs here
    funds_flow.xlsx      ← optional template (if --template flag is set)

Usage:
  python new_deal.py --deal "Project Nova" --closing-date 2026-06-30 --client-role buyer
  python new_deal.py --deal "Project Nova" --closing-date 2026-06-30 --client-role buyer \\
      --fund "Fund I=0.60" --fund "Fund II=0.40" --template

After running:
1. Drop your client's funds flow Excel into deals/<deal_slug>/
2. Drop all support documents into deals/<deal_slug>/documents/
3. Run the agent:
     python -m agent.main \\
       --deal "Project Nova" \\
       --closing-date 2026-06-30 \\
       --client-role buyer \\
       --funds-flow deals/project_nova/funds_flow.xlsx \\
       --documents deals/project_nova/documents/ \\
       --output deals/project_nova/run_output/
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ── Style helpers (matching build_funds_flow.py palette) ──────────────────────
def _S(s="thin"):
    return Side(style=s)

def _bdr():
    s = _S()
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(c):
    return PatternFill("solid", fgColor=c)

def _F(bold=False, size=10, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def _AL(h="left", v="center", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)


NAVY   = "1F3864"
WHITE  = "FFFFFF"
SUBHDR = "2E4057"
LBLUE  = "D6E4F0"


def slugify(name: str) -> str:
    """Convert deal name to filesystem-safe folder name."""
    s = name.lower().strip()
    s = re.sub(r"[^\w\s-]", "", s)
    s = re.sub(r"[\s_-]+", "_", s)
    return s


def scaffold(
    deal_name: str,
    closing_date: str,
    client_role: str,
    fund_allocations: dict,
    build_template: bool,
) -> Path:
    slug = slugify(deal_name)
    deal_dir = Path("deals") / slug

    if deal_dir.exists():
        print(f"  ⚠️  Folder already exists: {deal_dir}")
        print("     Skipping folder creation — existing files preserved.")
    else:
        deal_dir.mkdir(parents=True)
        print(f"  ✓  Created: {deal_dir}/")

    docs_dir   = deal_dir / "documents"
    output_dir = deal_dir / "run_output"
    docs_dir.mkdir(exist_ok=True)
    output_dir.mkdir(exist_ok=True)
    print(f"  ✓  Created: {docs_dir}/")
    print(f"  ✓  Created: {output_dir}/")

    if build_template:
        tmpl_path = deal_dir / "funds_flow_template.xlsx"
        _build_template(tmpl_path, deal_name, closing_date, client_role, fund_allocations)
        print(f"  ✓  Created: {tmpl_path}")
        print(f"     → Fill in line items and save as funds_flow.xlsx before running the agent.")

    # Write a README stub
    readme = deal_dir / "README.txt"
    if not readme.exists():
        readme.write_text(
            f"DEAL: {deal_name}\n"
            f"CLOSING DATE: {closing_date}\n"
            f"CLIENT ROLE: {client_role.upper()}\n"
            f"FUND ALLOCATIONS: {fund_allocations or 'not set'}\n\n"
            f"NEXT STEPS:\n"
            f"  1. Copy the client's funds flow Excel to this folder as: funds_flow.xlsx\n"
            f"  2. Drop all support documents (PDFs) into: documents/\n"
            f"  3. Run the agent:\n\n"
            f"     python -m agent.main \\\n"
            f"       --deal \"{deal_name}\" \\\n"
            f"       --closing-date {closing_date} \\\n"
            f"       --client-role {client_role} \\\n"
            f"       --funds-flow deals/{slug}/funds_flow.xlsx \\\n"
            f"       --documents deals/{slug}/documents/ \\\n"
            f"       --output deals/{slug}/run_output/\n"
        )
        print(f"  ✓  Created: {readme}")

    return deal_dir


def _build_template(
    path: Path,
    deal_name: str,
    closing_date: str,
    client_role: str,
    fund_allocations: dict,
) -> None:
    """Generate a minimal funds flow template Excel with the standard layout."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sources & Uses"

    # Column widths
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 36

    def cell(r, c, val=None, fnt=None, fll=None, aln=None, fmt=None):
        cl = ws.cell(row=r, column=c, value=val)
        if fnt: cl.font = fnt
        if fll: cl.fill = fll
        if aln: cl.alignment = aln
        if fmt: cl.number_format = fmt
        cl.border = _bdr()
        return cl

    def merge_row(r, label, fnt, fll, height=20):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        c = ws.cell(row=r, column=1, value=label)
        c.font = fnt; c.fill = fll
        c.alignment = _AL(indent=1)
        c.border = _bdr()
        ws.row_dimensions[r].height = height

    # Title
    merge_row(1, f"{deal_name.upper()} — SOURCES & USES  |  Closing: {closing_date}",
              _F(bold=True, size=13, color=WHITE), _fill(NAVY), height=30)

    # Column headers
    for ci, h in enumerate(["Description", "Fund I ($)", "Fund II ($)", "Total ($)", "Notes / Vendor"], 1):
        cell(2, ci, h, fnt=_F(bold=True, color=WHITE), fll=_fill(NAVY), aln=_AL("center"))
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    # Sections
    merge_row(3, "SOURCES", _F(bold=True, size=11, color=WHITE), _fill(SUBHDR))
    for r, label in [(4, "Capital Call — Fund I"), (5, "Capital Call — Fund II")]:
        cell(r, 1, f"  {label}", fnt=_F(), aln=_AL(indent=2))
        for c in [2, 3, 4]:
            cell(r, c, None, fnt=_F(), aln=_AL("center"), fmt='"$"#,##0')
        cell(r, 5, "", fnt=_F())
        ws.row_dimensions[r].height = 16
    cell(6, 1, "Total Sources", fnt=_F(bold=True), fll=_fill(LBLUE), aln=_AL(indent=1))
    for ci, col in enumerate("BCD", 2):
        cell(6, ci, f"={col}4+{col}5", fnt=_F(bold=True), fll=_fill(LBLUE),
             aln=_AL("center"), fmt='"$"#,##0')
    cell(6, 5, fll=_fill(LBLUE))
    ws.row_dimensions[6].height = 18

    ws.row_dimensions[7].height = 6

    merge_row(8, "USES", _F(bold=True, size=11, color=WHITE), _fill(SUBHDR))
    cell(9, 1, "  Investment in Target Co", fnt=_F(), aln=_AL(indent=2))
    for c in [2, 3, 4]:
        cell(9, c, None, fnt=_F(), aln=_AL("center"), fmt='"$"#,##0')
    cell(9, 5, "Per SPA", fnt=_F(italic=True))
    ws.row_dimensions[9].height = 16

    merge_row(10, f"TRANSACTION COSTS  ({client_role.upper()})",
              _F(bold=True, size=10, color=WHITE), _fill("344E6B"), height=16)

    # Placeholder cost lines (10 rows)
    for i in range(11, 21):
        cell(i, 1, f"  [Line item {i-10}]", fnt=_F(color="888888"), aln=_AL(indent=3))
        for c in [2, 3, 4]:
            cell(i, c, None, fnt=_F(), aln=_AL("center"), fmt='"$"#,##0')
        cell(i, 5, "[Vendor / Note]", fnt=_F(italic=True, color="888888"))
        ws.row_dimensions[i].height = 16

    cell(21, 1, "Total Transaction Costs", fnt=_F(bold=True), fll=_fill(LBLUE), aln=_AL(indent=2))
    for ci, col in enumerate("BCD", 2):
        cell(21, ci, f"=SUM({col}11:{col}20)", fnt=_F(bold=True), fll=_fill(LBLUE),
             aln=_AL("center"), fmt='"$"#,##0')
    cell(21, 5, fll=_fill(LBLUE))
    ws.row_dimensions[21].height = 18

    cell(22, 1, "Total Uses", fnt=_F(bold=True, size=11), fll=_fill(LBLUE), aln=_AL(indent=1))
    for ci, col in enumerate("BCD", 2):
        cell(22, ci, f"={col}9+{col}21", fnt=_F(bold=True, size=11), fll=_fill(LBLUE),
             aln=_AL("center"), fmt='"$"#,##0')
    cell(22, 5, fll=_fill(LBLUE))
    ws.row_dimensions[22].height = 20

    ws.row_dimensions[23].height = 6

    # Reconciliation
    merge_row(24, "RECONCILIATION", _F(bold=True, size=10, color=WHITE), _fill(SUBHDR), height=16)
    cell(25, 1, "Sources less Uses (must equal $0)", fnt=_F(italic=True), aln=_AL(indent=2))
    for ci, col in enumerate("BCD", 2):
        cell(25, ci, f"={col}6-{col}22", fnt=_F(bold=True, color="C00000"),
             aln=_AL("center"), fmt='"$"#,##0')
    cell(25, 5, "Must equal $0", fnt=_F(italic=True, color="555555"))
    ws.row_dimensions[25].height = 16

    # Wire tab stub
    ws2 = wb.create_sheet("Wire Instructions")
    ws2["A1"].value = f"Wire Transfer Instructions — {deal_name}"
    ws2["A1"].font = _F(bold=True, size=12)
    ws2["A3"].value = "Receiving Bank: [BANK NAME]"
    ws2["A4"].value = "ABA Routing: [ROUTING NUMBER]"
    ws2["A5"].value = "Account Name: [FUND NAME]"
    ws2["A6"].value = "Account Number: [ACCOUNT NUMBER]  (on file)"
    ws2["A8"].value = "Confirm wire details directly with Fund Finance before sending."
    ws2["A8"].font = _F(italic=True, color="C00000")
    ws2.column_dimensions["A"].width = 55

    wb.save(str(path))


def _parse_args():
    p = argparse.ArgumentParser(description="Scaffold a new deal folder for the Funds Flow Audit Tool.")
    p.add_argument("--deal",          required=True,  help="Deal name (e.g. 'Project Nova')")
    p.add_argument("--closing-date",  required=True,  help="Closing date YYYY-MM-DD")
    p.add_argument("--client-role",   required=True,  choices=["buyer", "seller", "both"])
    p.add_argument("--fund",          action="append", default=[],
                   help="Fund allocation e.g. 'Fund I=0.60' (repeat for each fund)")
    p.add_argument("--template",      action="store_true",
                   help="Generate a blank funds flow template Excel")
    return p.parse_args()


def _parse_allocations(fund_args: list[str]) -> dict:
    allocs = {}
    for arg in fund_args:
        if "=" in arg:
            name, pct = arg.split("=", 1)
            try:
                allocs[name.strip()] = float(pct.strip())
            except ValueError:
                pass
    return allocs


def main():
    args = _parse_args()
    allocs = _parse_allocations(args.fund)

    print(f"\nScaffolding deal: {args.deal}")
    print(f"  Closing: {args.closing_date}  |  Role: {args.client_role.upper()}")
    if allocs:
        print(f"  Allocations: {allocs}")
    print()

    deal_dir = scaffold(
        deal_name=args.deal,
        closing_date=args.closing_date,
        client_role=args.client_role,
        fund_allocations=allocs,
        build_template=args.template,
    )

    slug = slugify(args.deal)
    print(f"\nDeal folder ready: {deal_dir}/")
    print(f"\nNext steps:")
    print(f"  1. Copy client funds flow Excel into: {deal_dir}/")
    print(f"     (rename to funds_flow.xlsx)")
    print(f"  2. Copy all support documents into: {deal_dir}/documents/")
    print(f"  3. Run the agent:\n")
    print(f"     python -m agent.main \\")
    print(f'       --deal "{args.deal}" \\')
    print(f"       --closing-date {args.closing_date} \\")
    print(f"       --client-role {args.client_role} \\")
    print(f"       --funds-flow deals/{slug}/funds_flow.xlsx \\")
    print(f"       --documents deals/{slug}/documents/ \\")
    print(f"       --output deals/{slug}/run_output/")
    print()


if __name__ == "__main__":
    main()
