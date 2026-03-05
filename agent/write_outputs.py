"""
agent/write_outputs.py — Write final outputs from index.json.

Creates:
  - run_output/funds_flow_indexed.xlsx  — annotated workpaper with:
      • Match-status annotation columns on each deal tab
      • GL Account + Fund I/II formula columns + Σ Check (linked to Total column)
      • "Journal Entry" tab in standard format with formulas referencing deal tab cells
      • PDF snapshot tabs (FF01 · Vendor, etc.)
  - run_output/documents_indexed/       — FF-numbered copies of matched docs
  - run_output/documents_indexed/UNMATCHED/  — documents not tied to any line item

Usage:
  python agent/write_outputs.py <deal_dir>
"""
from __future__ import annotations

import io
import json
import re
import shutil
import sys
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Styles ────────────────────────────────────────────────────────────────────

STATUS_COLORS = {
    "MATCHED":    "C6EFCE",
    "PARTIAL":    "FFEB9C",
    "CUMULATIVE": "DDEBF7",
    "MISSING":    "FFC7CE",
    "UNMATCHED":  "E2EFDA",
}

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
CELL_FONT   = Font(name="Calibri", size=10)
BOLD_FONT   = Font(name="Calibri", bold=True, size=10)

THIN_SIDE   = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
MED_BOTTOM  = Border(bottom=Side(style="medium", color="1F3864"))

USD_FMT  = '"$"#,##0.00'
CHK_FMT  = '"$"#,##0.00_);[Red]("$"#,##0.00)'   # red parentheses for negatives


# ── Credit account codes ──────────────────────────────────────────────────────
# Loaded from chart_of_accounts.json. The LLM classifies debit-side GL accounts
# during the matching step and writes them into index.json — no rules here.

_COA_PATH = Path(__file__).parent.parent / "chart_of_accounts.json"

_DEFAULT_CREDIT = {
    "wire_payable":        ("2100", "Wire Payable"),
    "accrued_liabilities": ("2200", "Accrued Liabilities — Deal Costs"),
}


def _load_credit_accounts() -> dict:
    """Load credit-side account codes from chart_of_accounts.json."""
    if not _COA_PATH.exists():
        return _DEFAULT_CREDIT
    try:
        coa = json.loads(_COA_PATH.read_text())
        credits = {
            k: (v["code"], v["name"])
            for k, v in coa.get("credit_accounts", {}).items()
        }
        return credits or _DEFAULT_CREDIT
    except Exception as e:
        print(f"  Warning: could not load chart_of_accounts.json ({e}) — using defaults")
        return _DEFAULT_CREDIT


CREDIT_ACCOUNTS = _load_credit_accounts()


def _get_gl(item: dict) -> tuple[str, str]:
    """Return (gl_account_code, gl_account_name) from the item, with fallback."""
    code = item.get("gl_account_code") or "7120"
    name = item.get("gl_account_name") or "Transaction Costs — Miscellaneous Closing Costs"
    return code, name


# ── FF number assignment ──────────────────────────────────────────────────────

def _assign_ff_numbers(index: dict) -> None:
    """Assign FF refs sequentially; MISSING items get None. Mutates index in place."""
    ff_num = 1
    for item in index["line_items"]:
        if item.get("status") != "MISSING":
            item["ff_ref"] = f"FF{ff_num:02d}"
        else:
            item["ff_ref"] = None
        ff_num += 1


# ── Total column detector ─────────────────────────────────────────────────────

def _detect_total_col(ws) -> int | None:
    """Return 1-based column number of the Total/Amount column, or None."""
    for row in ws.iter_rows(max_row=6, values_only=True):
        for i, cell in enumerate(row, start=1):
            if cell and "total" in str(cell).lower():
                return i
    return None


# ── Annotate deal tabs + collect JE row info ──────────────────────────────────

def _write_annotated_excel(deal_dir: Path, index: dict) -> tuple[openpyxl.Workbook, Path, list[dict]]:
    """
    Copy funds_flow.xlsx → funds_flow_indexed.xlsx, add annotation columns.

    Returns (wb, out_path, je_row_info).  The caller is responsible for adding
    further sheets (JE tab, snapshots) and calling wb.save(out_path).
    """
    ff_path  = deal_dir / "funds_flow.xlsx"
    out_path = deal_dir / "run_output" / "funds_flow_indexed.xlsx"

    shutil.copy(str(ff_path), str(out_path))
    wb = openpyxl.load_workbook(str(out_path))

    lookup = {
        item["description"].strip().lower(): item
        for item in index["line_items"]
    }

    fi_pct  = index["fund_allocations"].get("Fund I",  0.55)
    fii_pct = index["fund_allocations"].get("Fund II", 0.45)

    # je_row_info: one dict per line item, keyed in index order
    # populated when we find the item's row in the worksheet
    je_row_info: list[dict] = []
    desc_to_je: dict[str, dict] = {}   # description_lower → je_row dict

    for item in index["line_items"]:
        key = item["description"].strip().lower()
        rec = {
            "ff_ref":      item.get("ff_ref"),
            "status":      item.get("status", "MISSING"),
            "acct_code":   "",
            "acct_name":   "",
            "description": item["description"],
            "sheet":       None,      # filled below
            "row":         None,      # filled below
            "fi_col":      None,      # column letter of Fund I formula cell
            "fii_col":     None,      # column letter of Fund II formula cell
            "ff_amount":   item["funds_flow_amount"],
            "doc_amount":  item.get("document_amount") or 0.0,
            "vendor":      item.get("document_vendor") or item["description"],
        }
        acct_code, acct_name = _get_gl(item)
        rec["acct_code"] = acct_code
        rec["acct_name"] = acct_name
        je_row_info.append(rec)
        desc_to_je[key] = rec

    for ws in wb.worksheets:
        title_lower = ws.title.lower()
        if any(k in title_lower for k in ("source", "wire", "instruction")):
            continue

        total_col_1 = _detect_total_col(ws)   # 1-based; may be None

        max_col = ws.max_column or 5
        ann     = max_col + 2   # start of annotation block (1-based), gap at max_col+1

        # ── Annotation column headers ─────────────────────────────────────────
        headers = [
            "FF Ref", "Match Status", "Document", "Amount Match", "Notes",
            # gap at ann+5
            "GL Account", f"Fund I ({fi_pct*100:.0f}%)", f"Fund II ({fii_pct*100:.0f}%)", "Σ Check",
        ]
        col_offsets = [0, 1, 2, 3, 4,  6, 7, 8, 9]   # offset 5 = blank gap

        for offset, h in zip(col_offsets, headers):
            c = ws.cell(1, ann + offset, h)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Column widths for annotation block
        col_widths = {0: 7, 1: 12, 2: 40, 3: 12, 4: 45, 6: 36, 7: 14, 8: 14, 9: 12}
        for offset, w in col_widths.items():
            ws.column_dimensions[get_column_letter(ann + offset)].width = w

        # ── Data rows ─────────────────────────────────────────────────────────
        for xl_row in ws.iter_rows(min_row=2):
            desc_cell = xl_row[0]
            if not desc_cell.value:
                continue
            key = str(desc_cell.value).strip().lower()
            if key not in lookup:
                continue

            item   = lookup[key]
            status = (item.get("status") or "").upper()
            color  = STATUS_COLORS.get(status, "FFFFFF")
            fill   = PatternFill("solid", fgColor=color)
            r      = desc_cell.row

            ff_ref_val = item.get("ff_ref") or "—"

            # Standard 5 annotation columns
            std_vals = [
                (ann,     ff_ref_val),
                (ann + 1, status),
                (ann + 2, item.get("document_file") or ""),
                (ann + 3, "YES" if item.get("amount_agrees") is True
                          else "NO" if item.get("amount_agrees") is False
                          else "N/A"),
                (ann + 4, item.get("notes") or ""),
            ]
            for col, val in std_vals:
                c = ws.cell(r, col, val)
                c.fill = fill
                c.font = CELL_FONT
                c.alignment = Alignment(vertical="center", wrap_text=(col == ann + 4))

            # GL Account (ann+6)
            acct_code, acct_name = _get_gl(item)
            gl_cell = ws.cell(r, ann + 6, f"{acct_code}  {acct_name}")
            gl_cell.fill = fill
            gl_cell.font = CELL_FONT
            gl_cell.alignment = Alignment(vertical="center")

            # Fund I / Fund II formula columns (ann+7, ann+8)
            if total_col_1:
                total_letter = get_column_letter(total_col_1)
                fi_letter    = get_column_letter(ann + 7)
                fii_letter   = get_column_letter(ann + 8)
                chk_letter   = get_column_letter(ann + 9)

                fi_cell = ws.cell(r, ann + 7, f"={total_letter}{r}*{fi_pct}")
                fi_cell.fill = fill
                fi_cell.font = CELL_FONT
                fi_cell.number_format = USD_FMT
                fi_cell.alignment = Alignment(horizontal="right", vertical="center")

                fii_cell = ws.cell(r, ann + 8, f"={total_letter}{r}*{fii_pct}")
                fii_cell.fill = fill
                fii_cell.font = CELL_FONT
                fii_cell.number_format = USD_FMT
                fii_cell.alignment = Alignment(horizontal="right", vertical="center")

                chk_cell = ws.cell(r, ann + 9,
                                   f"={fi_letter}{r}+{fii_letter}{r}-{total_letter}{r}")
                chk_cell.fill = fill
                chk_cell.font = CELL_FONT
                chk_cell.number_format = CHK_FMT
                chk_cell.alignment = Alignment(horizontal="right", vertical="center")

                # Store cell info for JE tab formula construction
                if key in desc_to_je:
                    rec = desc_to_je[key]
                    rec["sheet"]   = ws.title
                    rec["row"]     = r
                    rec["fi_col"]  = fi_letter
                    rec["fii_col"] = fii_letter

    return wb, out_path, je_row_info


# ── Journal Entry tab ─────────────────────────────────────────────────────────

def _cell_ref(sheet: str, col: str, row: int) -> str:
    """Return a single-quoted sheet cell reference for use in Excel formulas."""
    return f"'{sheet}'!{col}{row}"


def _build_je_tab(wb, index: dict, je_row_info: list[dict]) -> None:
    """
    Insert a 'Journal Entry' tab with two fully separate journal entries —
    one for Fund I and one for Fund II — each with their own debits, credits,
    and balance check. Amounts are formula-linked to the Buyer Expenses tab.

    Columns: Entry ID | Date | Account Code | Account Name | Description |
             FF Ref | Debit | Credit | Entry Type
    """
    closing_date = index["closing_date"]
    deal         = index["deal"]

    ws = wb.create_sheet("Journal Entry")

    def hcell(r, c, val):
        cell = ws.cell(r, c, val)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        return cell

    def dcell(r, c, val, *, bold=False, align="left", fill=None, fmt=None):
        cell = ws.cell(r, c, val)
        cell.font = BOLD_FONT if bold else CELL_FONT
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
        if fmt:
            cell.number_format = fmt
        return cell

    # 9 columns (no Fund column — each section is its own fund)
    COL_HEADERS = ["Entry ID", "Date", "Account Code", "Account Name",
                   "Description", "FF Ref", "Debit", "Credit", "Entry Type"]
    col_widths   = [10, 12, 14, 42, 55, 8, 14, 14, 10]
    for c, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # ── Title banner ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    title = ws["A1"]
    title.value = (
        f"JOURNAL ENTRY  |  {deal}  |  Closing: {closing_date}  |  "
        "Fund I: FF-JE-001-FI   Fund II: FF-JE-001-FII"
    )
    title.font  = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    title.fill  = HEADER_FILL
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:I2")
    note = ws["A2"]
    note.value = (
        "⚠  DRAFT — Debit/Credit amounts are formula-linked to 'Buyer Expenses' tab. "
        "Review with fund accounting before posting. Fund I and Fund II post separately."
    )
    note.font = Font(name="Calibri", italic=True, color="C00000", size=10)
    note.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # ── Pre-compute credit formula parts (shared across both fund sections) ──
    def _get_ref(rec: dict, col_key: str) -> str:
        """Return a cell reference string or raw value fallback."""
        if rec["sheet"] and rec[col_key]:
            return _cell_ref(rec["sheet"], rec[col_key], rec["row"])
        # fallback: look up hardcoded value from index
        item_obj = next(
            (i for i in index["line_items"] if i["description"] == rec["description"]), None)
        if item_obj:
            return str(item_obj["fund_i_amount"] if col_key == "fi_col" else item_obj["fund_ii_amount"])
        return "0"

    # Wire Payable parts and Accrued Liabilities parts, keyed by fund_col
    wp_parts: dict[str, list[str]] = {"fi_col": [], "fii_col": []}
    ac_parts: dict[str, list[str]] = {"fi_col": [], "fii_col": []}
    wp_descs: list[str] = []
    ac_descs: list[str] = []

    for rec in je_row_info:
        fi_ref  = _get_ref(rec, "fi_col")
        fii_ref = _get_ref(rec, "fii_col")
        ff_ref  = rec["ff_ref"] or "—"
        status  = rec["status"]
        ff_amt  = rec["ff_amount"]
        doc_amt = rec["doc_amount"]

        if status in ("MATCHED", "CUMULATIVE"):
            wp_parts["fi_col"].append(fi_ref)
            wp_parts["fii_col"].append(fii_ref)
            wp_descs.append(f"{ff_ref} {rec['description']}")
        elif status == "PARTIAL" and ff_amt:
            shortfall = ff_amt - doc_amt
            wp_parts["fi_col"].append(f"{fi_ref}*{doc_amt}/{ff_amt}")
            wp_parts["fii_col"].append(f"{fii_ref}*{doc_amt}/{ff_amt}")
            ac_parts["fi_col"].append(f"{fi_ref}*{shortfall}/{ff_amt}")
            ac_parts["fii_col"].append(f"{fii_ref}*{shortfall}/{ff_amt}")
            wp_descs.append(f"{ff_ref} {rec['description']} (${doc_amt:,.0f} of ${ff_amt:,.0f})")
            ac_descs.append(f"{ff_ref} {rec['description']} (shortfall ${shortfall:,.0f})")
        elif status == "MISSING":
            ac_parts["fi_col"].append(fi_ref)
            ac_parts["fii_col"].append(fii_ref)
            ac_descs.append(f"— {rec['description']} (no invoice)")

    def _cr_formula(parts: list[str]) -> str | int:
        return ("=" + "+".join(parts)) if parts else 0

    credit_fill_wp = PatternFill("solid", fgColor="C6EFCE")
    credit_fill_ac = PatternFill("solid", fgColor="FFEB9C")

    # ── Helper: write one complete fund JE section ────────────────────────────
    def _write_fund_section(start_row: int, fund_label: str,
                            je_ref: str, col_key: str) -> int:
        """Write debits + credits + balance check for one fund. Returns next row."""
        r = start_row

        # Fund section banner
        ws.merge_cells(f"A{r}:I{r}")
        banner = ws[f"A{r}"]
        banner.value = f"{fund_label}  —  {je_ref}"
        banner.font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        banner.fill  = PatternFill("solid", fgColor="2E4A7A")
        banner.alignment = Alignment(horizontal="left", indent=1, vertical="center")
        ws.row_dimensions[r].height = 20
        r += 1

        # Debit section label
        ws.merge_cells(f"A{r}:I{r}")
        dl = ws[f"A{r}"]
        dl.value = "DEBIT LINES"
        dl.font  = BOLD_FONT
        dl.fill  = PatternFill("solid", fgColor="DDEBF7")
        dl.alignment = Alignment(horizontal="left", indent=1, vertical="center")
        r += 1

        # Column headers
        for c, h in enumerate(COL_HEADERS, start=1):
            hcell(r, c, h)
        ws.row_dimensions[r].height = 18
        r += 1

        # Debit rows
        debit_rows: list[int] = []
        for rec in je_row_info:
            ref      = _get_ref(rec, col_key)
            formula  = f"={ref}"
            ff_ref   = rec["ff_ref"] or "—"
            desc_str = f"{ff_ref}  {rec['description']}  —  {rec['vendor']}"
            status   = rec["status"]
            fill     = PatternFill("solid", fgColor=STATUS_COLORS.get(status, "FFFFFF"))

            dcell(r, 1, je_ref,            align="center")
            dcell(r, 2, closing_date,      align="center")
            dcell(r, 3, rec["acct_code"],  align="center")
            dcell(r, 4, rec["acct_name"])
            dcell(r, 5, desc_str)
            dcell(r, 6, ff_ref,            align="center", fill=fill)
            dcell(r, 7, formula,           align="right",  fmt=USD_FMT)  # Debit
            dcell(r, 8, 0,                 align="right",  fmt=USD_FMT)  # Credit = 0
            dcell(r, 9, "debit",           align="center")
            debit_rows.append(r)
            r += 1

        # Total debits
        dr_sum_f = "=SUM(" + ",".join(f"G{dr}" for dr in debit_rows) + ")"
        ws.cell(r, 6, "Total Debits").font = BOLD_FONT
        ws.cell(r, 6).alignment = Alignment(horizontal="right")
        td = ws.cell(r, 7, dr_sum_f)
        td.font = BOLD_FONT; td.number_format = USD_FMT
        td.alignment = Alignment(horizontal="right"); td.border = MED_BOTTOM
        r += 2

        # Credit section label
        ws.merge_cells(f"A{r}:I{r}")
        cl = ws[f"A{r}"]
        cl.value = "CREDIT LINES"
        cl.font  = BOLD_FONT
        cl.fill  = PatternFill("solid", fgColor="C6EFCE")
        cl.alignment = Alignment(horizontal="left", indent=1, vertical="center")
        r += 1

        # Column headers
        for c, h in enumerate(COL_HEADERS, start=1):
            hcell(r, c, h)
        ws.row_dimensions[r].height = 18
        r += 1

        # Credit rows
        credit_rows: list[int] = []
        wp_f   = _cr_formula(wp_parts[col_key])
        ac_f   = _cr_formula(ac_parts[col_key])
        wp_basis = "; ".join(wp_descs[:3]) + ("…" if len(wp_descs) > 3 else "")

        wp_code, wp_name = CREDIT_ACCOUNTS.get("wire_payable",        ("2100", "Wire Payable"))
        ac_code, ac_name = CREDIT_ACCOUNTS.get("accrued_liabilities", ("2200", "Accrued Liabilities"))

        for acct_name, acct_code, cr_formula, fill, basis in [
            (wp_name, wp_code, wp_f, credit_fill_wp, wp_basis),
            (ac_name, ac_code, ac_f, credit_fill_ac, "; ".join(ac_descs)),
        ]:
            dcell(r, 1, je_ref,     align="center")
            dcell(r, 2, closing_date, align="center")
            dcell(r, 3, acct_code,  align="center")
            dcell(r, 4, acct_name,  bold=True, fill=fill)
            dcell(r, 5, basis,      fill=fill)
            dcell(r, 6, "",         fill=fill)
            dcell(r, 7, 0,          align="right", fmt=USD_FMT)           # Debit = 0
            dcell(r, 8, cr_formula, align="right", fmt=USD_FMT, fill=fill) # Credit
            dcell(r, 9, "credit",   align="center")
            credit_rows.append(r)
            r += 1

        # Total credits
        cr_sum_f = "=SUM(" + ",".join(f"H{cr}" for cr in credit_rows) + ")"
        ws.cell(r, 6, "Total Credits").font = BOLD_FONT
        ws.cell(r, 6).alignment = Alignment(horizontal="right")
        tc = ws.cell(r, 9, cr_sum_f)
        tc.font = BOLD_FONT; tc.number_format = USD_FMT
        tc.alignment = Alignment(horizontal="right"); tc.border = MED_BOTTOM
        r += 2

        # Balance check
        dr_cells = ",".join(f"G{dr}" for dr in debit_rows)
        cr_cells = ",".join(f"H{cr}" for cr in credit_rows)
        ws.merge_cells(f"A{r}:F{r}")
        bc_label = ws[f"A{r}"]
        bc_label.value = f"Balance Check — {fund_label}  (Debits − Credits = 0)"
        bc_label.font = BOLD_FONT
        bc_label.alignment = Alignment(horizontal="right", vertical="center")
        bc = ws.cell(r, 7, f"=SUM({dr_cells})-SUM({cr_cells})")
        bc.font = BOLD_FONT; bc.number_format = CHK_FMT
        bc.alignment = Alignment(horizontal="right"); bc.border = THIN_BORDER
        r += 3   # blank gap before next section

        return r

    # ── Write Fund I then Fund II ─────────────────────────────────────────────
    next_row = _write_fund_section(4, "Fund I",  "FF-JE-001-FI",  "fi_col")
    _write_fund_section(next_row, "Fund II", "FF-JE-001-FII", "fii_col")

    ws.freeze_panes = "A4"


# ── PDF snapshot tabs ─────────────────────────────────────────────────────────

def _add_snapshot_tabs(wb, deal_dir: Path, index: dict) -> None:
    """Add one sheet per matched document containing a rendered image of the invoice."""
    try:
        from pdf2image import convert_from_path
        from PIL import Image as PILImage
    except ImportError:
        print("  pdf2image / Pillow not available — skipping snapshots")
        return

    docs_dir = deal_dir / "documents"
    seen: set[str] = set()

    for item in index["line_items"]:
        doc_file = item.get("document_file")
        ff_ref   = item.get("ff_ref")
        if not doc_file or not ff_ref or item.get("status") == "MISSING":
            continue

        pdf_path = docs_dir / doc_file
        if not pdf_path.exists():
            print(f"  snapshot: file not found — {doc_file}")
            continue

        if doc_file not in seen:
            vendor_short = re.sub(r'[/\\?*:\[\]]', '-',
                                  (item.get("document_vendor") or doc_file)[:20])
            tab_name = f"{ff_ref} · {vendor_short}"[:31]

            ws = wb.create_sheet(tab_name)
            ws.column_dimensions["A"].width = 108
            ws.sheet_view.showGridLines = False

            ws.merge_cells("A1:B1")
            t = ws.cell(row=1, column=1,
                        value=f"{ff_ref}  |  {item.get('document_vendor', '')}  |  {doc_file}")
            t.font      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            t.fill      = PatternFill("solid", fgColor="1F3864")
            t.alignment = Alignment(horizontal="left", indent=1)
            ws.row_dimensions[1].height = 22

            pages = convert_from_path(str(pdf_path), dpi=130, first_page=1, last_page=1)
            img   = pages[0]
            scale = 800 / img.width
            new_h = int(img.height * scale)
            img   = img.resize((800, new_h), PILImage.LANCZOS)

            buf = io.BytesIO()
            img.save(buf, format="PNG")
            buf.seek(0)

            xl_img        = XLImage(buf)
            xl_img.anchor = "A2"
            ws.add_image(xl_img)
            ws.row_dimensions[2].height = new_h * 0.74

            seen.add(doc_file)
            print(f"  snapshot: {tab_name}")


# ── Copy numbered documents ───────────────────────────────────────────────────

def _copy_numbered_docs(deal_dir: Path, index: dict) -> tuple[int, int]:
    docs_dir      = deal_dir / "documents"
    idx_dir       = deal_dir / "run_output" / "documents_indexed"
    unmatched_dir = idx_dir / "UNMATCHED"

    if idx_dir.exists():
        shutil.rmtree(idx_dir)
    idx_dir.mkdir(parents=True)
    unmatched_dir.mkdir()

    matched_files: set[str] = set()
    matched_count = 0

    for item in index["line_items"]:
        doc_file = item.get("document_file")
        ff_ref   = item.get("ff_ref")
        if not doc_file or not ff_ref or item.get("status") == "MISSING":
            continue
        src = docs_dir / doc_file
        if not src.exists():
            continue

        vendor    = (item.get("document_vendor") or "").replace("/", "-")[:30].strip()
        dest_name = f"{ff_ref} - {vendor} - {doc_file}"
        shutil.copy(str(src), str(idx_dir / dest_name))
        matched_files.add(doc_file)
        matched_count += 1
        print(f"  {ff_ref}: {doc_file}")

    unmatched_count = 0
    for pdf in sorted(docs_dir.glob("*.pdf")):
        if pdf.name not in matched_files:
            shutil.copy(str(pdf), str(unmatched_dir / pdf.name))
            print(f"  UNMATCHED: {pdf.name}")
            unmatched_count += 1

    return matched_count, unmatched_count


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: python agent/write_outputs.py <deal_dir>")
        sys.exit(1)

    deal_dir   = Path(sys.argv[1])
    index_path = deal_dir / "run_output" / "index.json"

    if not index_path.exists():
        print(f"ERROR: index.json not found at {index_path}")
        sys.exit(1)

    index = json.loads(index_path.read_text())

    # Assign FF numbers and persist back to index.json
    _assign_ff_numbers(index)
    index_path.write_text(json.dumps(index, indent=2))

    print("Writing annotated Excel...")
    wb, out_path, je_row_info = _write_annotated_excel(deal_dir, index)

    print("  Building Journal Entry tab...")
    _build_je_tab(wb, index, je_row_info)
    _add_snapshot_tabs(wb, deal_dir, index)
    wb.save(str(out_path))
    print(f"  → {out_path.name}")

    print("\nCopying numbered documents...")
    matched, unmatched = _copy_numbered_docs(deal_dir, index)
    print(f"  → {matched} matched, {unmatched} unmatched")

    print("\nDone.")


if __name__ == "__main__":
    main()
