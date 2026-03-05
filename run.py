"""
run.py — Funds Flow Audit Tool

1. Drop your funds flow Excel (.xlsx) and support documents (.pdf) into input/
2. Run:  python run.py
3. Files are moved to deals/<deal-name>/ and the agent runs there
4. input/ is empty when done — ready for the next deal

Output is written to deals/<deal-name>/run_output/
"""
from __future__ import annotations

import re
import shutil
import sys
from pathlib import Path

import openpyxl

INPUT_DIR   = Path("input")
DEALS_DIR   = Path("deals")

# Row/col positions of deal metadata in the Sources & Uses sheet
_R_INFO  = 2   # col 2=deal name, col 4=closing date, col 6=client role
_R_FUNDS = 3   # col 2=fund I name, col 4=fund I %, col 6=fund II name, col 8=fund II %


def find_excel() -> Path:
    files = [f for f in INPUT_DIR.glob("*.xlsx") if not f.name.startswith("~")]
    if not files:
        print("ERROR: No .xlsx file found in input/")
        print("       Drop your client's funds flow Excel into input/ and try again.")
        sys.exit(1)
    if len(files) > 1:
        files.sort(key=lambda f: f.stat().st_mtime, reverse=True)
        print(f"  Multiple .xlsx files found — using most recent: {files[0].name}")
    return files[0]


def read_metadata(excel_path: Path) -> dict:
    wb = openpyxl.load_workbook(str(excel_path), data_only=True, read_only=True)

    # Find the Sources & Uses sheet
    ws = next(
        (wb[n] for n in wb.sheetnames if "source" in n.lower() or "uses" in n.lower()),
        wb.worksheets[0]
    )

    def val(row, col):
        v = ws.cell(row, col).value
        return str(v).strip() if v is not None else ""

    def pct(row, col):
        v = ws.cell(row, col).value
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip().rstrip("%")
        try:
            f = float(s)
            return f / 100 if f > 1 else f
        except ValueError:
            return None

    meta = {
        "deal_name":        val(_R_INFO,  2),
        "closing_date":     val(_R_INFO,  4),
        "client_role":      val(_R_INFO,  6).lower() or "buyer",
        "fund_allocations": {},
    }

    f1_name, f1_pct = val(_R_FUNDS, 2), pct(_R_FUNDS, 4)
    f2_name, f2_pct = val(_R_FUNDS, 6), pct(_R_FUNDS, 8)
    if f1_name and f1_pct is not None:
        meta["fund_allocations"][f1_name] = f1_pct
    if f2_name and f2_pct is not None:
        meta["fund_allocations"][f2_name] = f2_pct

    wb.close()
    return meta


def slugify(name: str) -> str:
    s = name.lower().strip()
    s = re.sub(r"[^\w\s-]", "", s)
    return re.sub(r"[\s_-]+", "_", s)


def stage_files(excel: Path, meta: dict) -> tuple[Path, Path, Path]:
    """
    Move files from input/ into deals/<deal-name>/.
    Returns (deal_dir, funds_flow_dest, documents_dir).
    """
    deal_dir   = DEALS_DIR / slugify(meta["deal_name"])
    docs_dir   = deal_dir / "documents"
    output_dir = deal_dir / "run_output"

    deal_dir.mkdir(parents=True, exist_ok=True)
    docs_dir.mkdir(exist_ok=True)
    output_dir.mkdir(exist_ok=True)

    # Move the Excel
    ff_dest = deal_dir / "funds_flow.xlsx"
    shutil.move(str(excel), str(ff_dest))
    print(f"  Moved: {excel.name}  →  {ff_dest}")

    # Move all PDFs into documents/
    pdfs = list(INPUT_DIR.glob("*.pdf"))
    for pdf in pdfs:
        dest = docs_dir / pdf.name
        shutil.move(str(pdf), str(dest))
    print(f"  Moved: {len(pdfs)} PDF(s)  →  {docs_dir}/")

    return deal_dir, ff_dest, docs_dir, output_dir


def main():
    # Check input/ has files
    all_files = [f for f in INPUT_DIR.iterdir() if f.is_file()]
    if not all_files:
        print("input/ is empty. Drop your funds flow Excel and support PDFs there, then run again.")
        sys.exit(0)

    excel = find_excel()

    print(f"\n  Reading {excel.name}...")
    meta = read_metadata(excel)

    if not meta["deal_name"]:
        print("ERROR: Could not read deal name from the Excel header (row 2, column 2).")
        print("       Make sure the Excel uses the Funds Flow Audit Tool template.")
        sys.exit(1)

    pdf_count = len(list(INPUT_DIR.glob("*.pdf")))

    print()
    print("=" * 60)
    print("  FUNDS FLOW INDEXER")
    print("=" * 60)
    print(f"  Deal:        {meta['deal_name']}")
    print(f"  Closing:     {meta['closing_date']}   Role: {meta['client_role'].upper()}")
    print(f"  Excel:       {excel.name}")
    print(f"  Documents:   {pdf_count} PDF(s)")
    if meta["fund_allocations"]:
        alloc_str = "  /  ".join(f"{k}: {v:.0%}" for k, v in meta["fund_allocations"].items())
        print(f"  Allocation:  {alloc_str}")
    print("=" * 60)
    print()

    # Move files from input/ → deals/<deal>/
    print("  Staging files...")
    deal_dir, ff_dest, docs_dir, output_dir = stage_files(excel, meta)
    print(f"  Deal folder: {deal_dir}/")
    print()

    # Write last_run.json so the skill command knows where to find the files
    import json
    last_run = {
        "deal_name":        meta["deal_name"],
        "closing_date":     meta["closing_date"],
        "client_role":      meta["client_role"],
        "fund_allocations": meta["fund_allocations"],
        "deal_dir":         str(deal_dir),
        "funds_flow_path":  str(ff_dest),
        "documents_dir":    str(docs_dir),
        "output_dir":       str(output_dir),
    }
    Path("last_run.json").write_text(json.dumps(last_run, indent=2))
    print(f"  Ready. Run /index-funds-flow to index this deal.")


if __name__ == "__main__":
    main()
