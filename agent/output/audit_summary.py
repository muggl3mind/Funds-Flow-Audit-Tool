"""
Audit Summary sheet for the agent/main.py pipeline's annotated workpaper.

Renders a scoreboard, full index table, and exceptions list.
"""
from __future__ import annotations

from openpyxl.utils import get_column_letter

from agent.config import DealConfig
from agent.exceptions.exception_classifier import DealException
from agent.matcher.llm_matcher import MatchResult
from agent.normalizers.document_normalizer import DocumentRecord
from agent.normalizers.funds_flow_normalizer import TabScopeResult
from agent.utils.amount_utils import format_usd

# Import the style helpers and constants from excel_writer
from agent.output.excel_writer import (
    _font, _fill, _align, _border, STATUS_SYMBOL,
    C_NAVY, C_WHITE, C_GREEN_BG, C_GREEN_FG, C_RED_BG, C_RED_FG,
    C_AMBER_BG, C_AMBER_FG, C_BLUE_BG, C_GREY_BG, C_AUDIT_HDR,
    SEV_COLORS,
)


def add_audit_summary(
    wb,
    results: list[MatchResult],
    exceptions: list[DealException],
    scope: TabScopeResult,
    all_documents: list[DocumentRecord],
    config: DealConfig,
    summary: dict,
):
    # Remove existing audit summary if re-running
    if "Audit Summary" in wb.sheetnames:
        del wb["Audit Summary"]

    ws = wb.create_sheet("Audit Summary")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 40

    # Title
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"{config.deal_name} — Audit Support Index"
    c.font = _font(bold=True, size=13, color=C_WHITE)
    c.fill = _fill(C_NAVY)
    c.alignment = _align(h="center")
    ws.row_dimensions[1].height = 24

    # Deal metadata
    ws.merge_cells("A2:G2")
    ws["A2"].value = (
        f"Closing: {config.closing_date}  |  "
        f"Indexed: {summary.get('indexed_at', '')}  |  "
        f"Client role: {config.client_role.upper()}  |  "
        f"Agent v{config.agent_version}"
    )
    ws["A2"].font = _font(italic=True, size=9, color=C_WHITE)
    ws["A2"].fill = _fill(C_AUDIT_HDR)
    ws["A2"].alignment = _align(h="center")
    ws.row_dimensions[2].height = 14

    ws.row_dimensions[3].height = 6

    # Scoreboard
    score_items = [
        ("Total Line Items",        summary["total_line_items"],  None),
        ("Matched",                 summary["matched"],           C_GREEN_FG),
        ("Partial",                 summary["partial"],           C_AMBER_FG),
        ("Missing",                 summary["missing"],           C_RED_FG),
        ("Exceptions",              summary.get("exceptions", 0), C_AMBER_FG),
        ("Amount Mismatches",       summary["amount_mismatches"], C_RED_FG if summary["amount_mismatches"] else C_GREEN_FG),
        ("Total Supported ($)",     summary["total_supported_amount"],   None),
        ("Total Unsupported ($)",   summary["total_unsupported_amount"], C_RED_FG),
        ("Avg Match Confidence",    summary.get("match_confidence_avg"), None),
    ]
    for r_off, (label, value, fg) in enumerate(score_items, start=4):
        lc = ws.cell(r_off, 1, label)
        vc = ws.cell(r_off, 2, value)
        lc.font = _font(bold=True)
        vc.font = _font(bold=True, color=fg or "000000")
        for c in (lc, vc):
            c.fill = _fill(C_BLUE_BG)
            c.border = _border()
            c.alignment = _align(indent=1)
        if label in ("Total Supported ($)", "Total Unsupported ($)") and value:
            vc.number_format = '"$"#,##0'
        if label == "Avg Match Confidence" and value:
            vc.number_format = "0%"
        ws.row_dimensions[r_off].height = 16

    # Tabs in scope
    r = len(score_items) + 5
    ws.cell(r, 1, "Tabs In Scope").font = _font(bold=True)
    ws.cell(r, 2, ", ".join(scope.in_scope) or "(none)").font = _font(color=C_GREEN_FG)
    r += 1
    ws.cell(r, 1, "Tabs Skipped").font = _font(bold=True)
    ws.cell(r, 2, ", ".join(scope.skipped) or "(none)").font = _font(color=C_RED_FG)
    for row in (r - 1, r):
        ws.row_dimensions[row].height = 14

    r += 2

    # Full index table
    ws.row_dimensions[r].height = 6
    r += 1

    headers = ["Description", "Fund Allocations", "FF Amount",
               "Doc Amount", "Agrees?", "Status", "Document / Note"]
    hdrs_w  = [32, 20, 14, 14, 10, 12, 46]
    for col, (h, w) in enumerate(zip(headers, hdrs_w), 1):
        cell = ws.cell(r, col, h)
        cell.font = _font(bold=True, color=C_WHITE)
        cell.fill = _fill(C_NAVY)
        cell.alignment = _align(h="center")
        cell.border = _border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[r].height = 20
    r += 1

    for result in results:
        item = result.line_item
        doc  = result.matched_document
        bg   = (C_GREEN_BG  if result.status == "matched"  else
                C_AMBER_BG  if result.status in ("partial", "exception") else
                C_RED_BG)

        alloc_str = "  |  ".join(
            f"{k}: {format_usd(v)}" for k, v in item.fund_allocations.items()
        ) if item.fund_allocations else "—"
        agrees_str = {True: "✓ Yes", False: "✗ No", None: "—"}[result.amount_agrees]
        status_str = f"{STATUS_SYMBOL.get(result.status,'?')}  {result.status.upper()}"
        doc_str    = doc.file_name if doc else "— NO DOCUMENT —"

        row_data = [
            item.description, alloc_str,
            item.total_amount, doc.total_amount if doc else None,
            agrees_str, status_str, doc_str,
        ]
        status_fg = (C_GREEN_FG if result.status == "matched" else
                     C_RED_FG   if result.status == "missing" else "000000")

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(r, col, value)
            cell.fill = _fill(bg)
            cell.border = _border()
            cell.font = _font(
                bold=(col == 6),
                color=status_fg if col == 6 else "000000",
            )
            cell.alignment = _align(
                h="right"  if col in (3, 4) else
                "center" if col in (5, 6) else "left"
            )
            if col in (3, 4) and value is not None:
                cell.number_format = '"$"#,##0'
        ws.row_dimensions[r].height = 16
        r += 1

    # Exceptions
    r += 1
    ws.merge_cells(f"A{r}:G{r}")
    ws.cell(r, 1, "EXCEPTIONS & ACTION ITEMS").font = _font(bold=True, color=C_WHITE, size=11)
    ws.cell(r, 1).fill = _fill(C_NAVY)
    ws.cell(r, 1).alignment = _align(h="left", indent=1)
    ws.row_dimensions[r].height = 18
    r += 1

    exc_headers = ["Severity", "Type", "Source", "Description", "Suggested Action"]
    exc_widths  = [10, 24, 26, 44, 44]
    for col, (h, w) in enumerate(zip(exc_headers, exc_widths), 1):
        cell = ws.cell(r, col, h)
        cell.font = _font(bold=True, color=C_WHITE)
        cell.fill = _fill(C_AUDIT_HDR)
        cell.border = _border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[r].height = 18
    r += 1

    if not exceptions:
        ws.cell(r, 1, "No exceptions found.").font = _font(italic=True)
    else:
        for exc in exceptions:
            bg = SEV_COLORS.get(exc.severity, C_GREY_BG)
            for col, val in enumerate(
                [exc.severity, exc.exception_type, exc.source_ref,
                 exc.description, exc.suggested_action], 1
            ):
                cell = ws.cell(r, col, val)
                cell.fill = _fill(bg)
                cell.border = _border()
                cell.font = _font(bold=(col == 1))
                cell.alignment = _align(wrap=(col in (4, 5)), h="center" if col <= 2 else "left")
            ws.row_dimensions[r].height = 32
            r += 1

    ws.freeze_panes = "A4"
