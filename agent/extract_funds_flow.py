"""
agent/extract_funds_flow.py — Extract line items from a funds flow Excel.

Reads the workbook, skips non-scope tabs (seller, wire, sources & uses summary),
and returns structured JSON with every transaction cost line item.

Usage:
  python agent/extract_funds_flow.py <excel_path> <output_json>
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

SKIP_KEYWORDS  = {"seller", "wire", "instruction", "wiring", "bank"}
TOTAL_KEYWORDS = {"total", "subtotal", "grand total", "sources", "uses",
                  "reconcil", "investment", "capital call"}


def _to_num(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").replace("$", "").strip())
    except ValueError:
        return None


def _is_skip_tab(name: str) -> bool:
    n = name.lower()
    return any(k in n for k in SKIP_KEYWORDS) or "source" in n


def _extract_metadata(wb) -> dict:
    ws = next(
        (wb[n] for n in wb.sheetnames if "source" in n.lower() or "use" in n.lower()),
        wb.worksheets[0],
    )

    def val(r, c):
        v = ws.cell(r, c).value
        return str(v).strip() if v is not None else ""

    def pct(r, c):
        v = ws.cell(r, c).value
        if v is None:
            return None
        if isinstance(v, (int, float)):
            f = float(v)
            return f / 100 if f > 1 else f
        try:
            s = str(v).strip().rstrip("%")
            f = float(s)
            return f / 100 if f > 1 else f
        except ValueError:
            return None

    meta = {
        "deal":             val(2, 2),
        "closing_date":     val(2, 4),
        "client_role":      val(2, 6).lower() or "buyer",
        "fund_allocations": {},
    }
    f1n, f1p = val(3, 2), pct(3, 4)
    f2n, f2p = val(3, 6), pct(3, 8)
    if f1n and f1p is not None:
        meta["fund_allocations"][f1n] = f1p
    if f2n and f2p is not None:
        meta["fund_allocations"][f2n] = f2p
    return meta


def _detect_columns(ws) -> dict:
    """Scan header rows to find which column index holds Fund I, Fund II, Total, Notes."""
    cols = {"fund_i": None, "fund_ii": None, "total": None, "notes": None}
    for row in ws.iter_rows(max_row=5, values_only=True):
        for i, cell in enumerate(row):
            if cell is None:
                continue
            v = str(cell).lower()
            if "fund i" in v and "ii" not in v and cols["fund_i"] is None:
                cols["fund_i"] = i
            elif "fund ii" in v and cols["fund_ii"] is None:
                cols["fund_ii"] = i
            elif "total" in v and cols["total"] is None:
                cols["total"] = i
            elif ("note" in v or "vendor" in v or "ref" in v) and cols["notes"] is None:
                cols["notes"] = i
        if all(v is not None for v in cols.values()):
            break
    return cols


def _extract_tab(ws, tab_name: str) -> list[dict]:
    cols  = _detect_columns(ws)
    fi_c  = cols["fund_i"]
    fii_c = cols["fund_ii"]
    tot_c = cols["total"]
    not_c = cols["notes"]

    items = []
    for row in ws.iter_rows(values_only=False):
        row_idx = row[0].row

        desc_cell = row[0]
        desc = desc_cell.value
        if desc is None:
            continue
        desc_str = str(desc).strip()
        if not desc_str:
            continue

        # Skip section headers / totals
        dl = desc_str.lower()
        if any(k in dl for k in TOTAL_KEYWORDS):
            continue

        def _get(idx):
            return _to_num(row[idx].value) if idx is not None and len(row) > idx else None

        fund_i = _get(fi_c)
        fund_ii = _get(fii_c)
        total   = _get(tot_c)

        if all(a is None for a in (fund_i, fund_ii, total)):
            continue

        # Derive total if missing
        if total is None and fund_i is not None and fund_ii is not None:
            total = fund_i + fund_ii
        if total is None:
            total = next((a for a in (fund_i, fund_ii) if a is not None), None)

        if total is None or total <= 0:
            continue

        notes = ""
        if not_c is not None and len(row) > not_c and row[not_c].value:
            notes = str(row[not_c].value).strip()

        items.append({
            "tab":            tab_name,
            "row":            row_idx,
            "description":    desc_str,
            "fund_i_amount":  fund_i,
            "fund_ii_amount": fund_ii,
            "total_amount":   total,
            "notes":          notes,
        })
    return items


def main():
    if len(sys.argv) < 3:
        print("Usage: python agent/extract_funds_flow.py <excel_path> <output_json>")
        sys.exit(1)

    excel_path  = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    wb   = openpyxl.load_workbook(str(excel_path), data_only=True)
    meta = _extract_metadata(wb)

    all_items:  list[dict] = []
    processed:  list[str]  = []
    skipped:    list[str]  = []

    for name in wb.sheetnames:
        if _is_skip_tab(name):
            skipped.append(name)
            continue
        items = _extract_tab(wb[name], name)
        all_items.extend(items)
        processed.append(name)

    output = {
        **meta,
        "line_items":     all_items,
        "tabs_processed": processed,
        "tabs_skipped":   skipped,
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(output, indent=2))
    print(f"Extracted {len(all_items)} line items from {processed} → {output_path}")


if __name__ == "__main__":
    main()
